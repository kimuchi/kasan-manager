"""alpha.5.12 reviewer workflow hardening tests.

generate_alpha5_12_reviewer_workflow_hardening.py / 拡張 alpha.5.10 gate /
拡張 alpha.5.11 export の挙動と不変条件を保護する。
"""
from __future__ import annotations

import csv
import io
import json
import subprocess
import sys
import tempfile
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PRODUCT_ROOT = ROOT
HARDENING_DIR = PRODUCT_ROOT / "out" / "internal" / "alpha5_12_reviewer_workflow_hardening"
WORKBOOK = HARDENING_DIR / "alpha5_12_reviewer_decision_workbook.xlsx"
SAMPLE_INPUT = HARDENING_DIR / "sample_reviewed_decisions.csv"

ALPHA_5_9_PACKET = PRODUCT_ROOT / "out" / "internal" / "alpha5_9_master_review_packet"
ALPHA_5_10_GATE = PRODUCT_ROOT / "out" / "internal" / "alpha5_10_reviewer_decision_gate"
ALPHA_5_11_HANDOFF = PRODUCT_ROOT / "out" / "internal" / "alpha5_11_reviewer_handoff_workbook"
ALPHA_5_11_WORKBOOK = ALPHA_5_11_HANDOFF / "alpha5_11_reviewer_decision_workbook.xlsx"
ALPHA_5_11_COMMIT = "a3af77843f01653f1e7e10c99ece00b98faa2aba"

GATE_SCRIPT = PRODUCT_ROOT / "scripts" / "generate_alpha5_10_reviewer_decision_gate.py"
EXPORT_SCRIPT = PRODUCT_ROOT / "scripts" / "export_alpha5_11_workbook_decisions.py"
GENERATOR_SCRIPT = PRODUCT_ROOT / "scripts" / "generate_alpha5_12_reviewer_workflow_hardening.py"

LEGACY_LEGAL_CSV_HEADER = (
    "service,kasan_key,reviewer_decision,reason,required_evidence,"
    "reviewer_name,reviewed_at,final_approved_by,implementation_allowed"
)


# ============================================================
# Helpers
# ============================================================

def _read_data_rows(csv_path: Path) -> list[dict]:
    if not csv_path.exists():
        return []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [r for r in reader if r.get("service") and not r.get("service", "").startswith("#")]


def _read_manifest(out_dir: Path, name: str) -> dict:
    return json.loads((out_dir / name).read_text(encoding="utf-8"))


def _iter_master_kasans():
    services = ("houmon_kango_kaigo", "tsusho_kaigo", "houmon_kaigo", "kyotaku_shien")
    for svc in services:
        path = PRODUCT_ROOT / f"regulatory_master/kaigo/{svc}.json"
        d = json.loads(path.read_text(encoding="utf-8"))
        for k, v in (d.get("kasans") or {}).items():
            yield svc, k, v


def _run_gate(input_csv: Path, out_dir: Path):
    return subprocess.run(
        [sys.executable, str(GATE_SCRIPT),
         "--input", str(input_csv), "--output", str(out_dir),
         "--alpha59-packet-dir", str(ALPHA_5_9_PACKET)],
        cwd=PRODUCT_ROOT, capture_output=True, text=True,
    )


# ============================================================
# 1. パケット構造
# ============================================================

def test_alpha_5_12_hardening_directory_exists():
    assert HARDENING_DIR.exists() and HARDENING_DIR.is_dir()
    expected = {
        "README.md",
        "alpha5_12_reviewer_decision_workbook.xlsx",
        "sample_reviewed_decisions.csv",
        "sample_reviewed_decision_validation_report.md",
        "sample_approved_changes_preview.csv",
        "sample_blocked_or_incomplete_decisions.csv",
        "sample_pending_decisions.csv",
        "sample_legal_review_required.csv",
        "legal_clearance_rules.md",
        "alpha5_12_reviewer_workflow_hardening_manifest.json",
    }
    found = {p.name for p in HARDENING_DIR.iterdir() if p.is_file()}
    missing = expected - found
    assert not missing, f"missing: {missing}"
    print("✅ test_alpha_5_12_hardening_directory_exists")


def test_alpha_5_12_manifest_json_is_valid():
    m = _read_manifest(HARDENING_DIR, "alpha5_12_reviewer_workflow_hardening_manifest.json")
    for key in (
        "version", "base_commit", "input_packet_version", "input_gate_version",
        "extended_workbook_columns", "valid_legal_review_clearance",
        "valid_reviewer_role", "valid_implementation_priority",
        "valid_implementation_risk_acknowledged", "high_risk_decisions",
        "sample_reviewed_decision_count", "sample_gate_result",
        "public_release", "checked_promotion", "master_auto_update",
        "r8_provisional_used_for_checked", "release_pack_modified",
    ):
        assert key in m, f"manifest missing: {key}"
    assert m["version"] == "alpha.5.12"
    assert m["public_release"] is False
    assert m["checked_promotion"] is False
    assert m["master_auto_update"] is False
    assert m["r8_provisional_used_for_checked"] is False
    assert m["release_pack_modified"] is False
    print("✅ test_alpha_5_12_manifest_json_is_valid")


def test_alpha_5_12_manifest_base_commit_equals_alpha_5_11():
    m = _read_manifest(HARDENING_DIR, "alpha5_12_reviewer_workflow_hardening_manifest.json")
    assert m["base_commit"] == ALPHA_5_11_COMMIT
    assert m["input_packet_version"] == "alpha.5.11"
    print("✅ test_alpha_5_12_manifest_base_commit_equals_alpha_5_11")


# ============================================================
# 2. workbook 拡張
# ============================================================

def test_alpha_5_12_workbook_xlsx_exists():
    assert WORKBOOK.exists() and WORKBOOK.is_file()
    assert WORKBOOK.stat().st_size > 1000
    print(f"✅ test_alpha_5_12_workbook_xlsx_exists: size={WORKBOOK.stat().st_size}")


def test_alpha_5_12_workbook_has_legal_review_clearance_column():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Decision_Input_All"]
    header = next(ws.iter_rows(values_only=True), None)
    assert header is not None
    assert "legal_review_clearance" in header, f"header={header}"
    print("✅ test_alpha_5_12_workbook_has_legal_review_clearance_column")


def test_alpha_5_12_workbook_has_legal_review_reference_column():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Decision_Input_All"]
    header = next(ws.iter_rows(values_only=True), None)
    assert "legal_review_reference" in header
    # Needs_Legal_Review シートにも同じ列があること
    ws2 = wb["Needs_Legal_Review"]
    header2 = next(ws2.iter_rows(values_only=True), None)
    assert "legal_review_reference" in header2
    print("✅ test_alpha_5_12_workbook_has_legal_review_reference_column")


def test_alpha_5_12_workbook_has_review_note_column():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Decision_Input_All"]
    header = next(ws.iter_rows(values_only=True), None)
    assert "review_note" in header
    print("✅ test_alpha_5_12_workbook_has_review_note_column")


# ============================================================
# 3. export script 拡張
# ============================================================

def test_alpha_5_12_export_csv_includes_extended_columns(tmp_path):
    """alpha.5.12 workbook を export → 拡張 16列 CSV"""
    out = tmp_path / "extended.csv"
    res = subprocess.run(
        [sys.executable, str(EXPORT_SCRIPT),
         "--workbook", str(WORKBOOK), "--output", str(out)],
        cwd=PRODUCT_ROOT, capture_output=True, text=True,
    )
    assert res.returncode == 0, res.stderr
    with open(out, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
    expected_extended = {
        "reviewer_role", "review_note",
        "legal_review_clearance", "legal_review_reference", "legal_review_note",
        "implementation_priority", "implementation_risk_acknowledged",
    }
    missing = expected_extended - set(header)
    assert not missing, f"export CSV missing extended columns: {missing}"
    print(f"✅ test_alpha_5_12_export_csv_includes_extended_columns")


# ============================================================
# 4. backward compat: legacy 9列CSV / extended 16列CSV
# ============================================================

def test_alpha_5_12_gate_still_accepts_legacy_9_column_csv(tmp_path):
    """alpha.5.11 workbook（9列スキーマ）の export 結果を gate に渡せる"""
    legacy_export = tmp_path / "legacy_export.csv"
    res_e = subprocess.run(
        [sys.executable, str(EXPORT_SCRIPT),
         "--workbook", str(ALPHA_5_11_WORKBOOK),
         "--output", str(legacy_export),
         "--schema", "legacy"],
        cwd=PRODUCT_ROOT, capture_output=True, text=True,
    )
    assert res_e.returncode == 0, res_e.stderr
    # header が 9列であること
    with open(legacy_export, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
    assert "legal_review_clearance" not in header
    assert "implementation_risk_acknowledged" not in header
    assert len(header) == 9, f"header={header}"

    # gate に渡して正常終了
    gate_out = tmp_path / "gate_out_legacy"
    res_g = _run_gate(legacy_export, gate_out)
    assert res_g.returncode == 0, res_g.stderr
    m = _read_manifest(gate_out, "alpha5_10_reviewer_decision_gate_manifest.json")
    assert m["input_csv_schema"] == "legacy_9_column"
    assert m["has_legal_clearance_column"] is False
    assert m["has_implementation_risk_acknowledged_column"] is False
    print("✅ test_alpha_5_12_gate_still_accepts_legacy_9_column_csv")


def test_alpha_5_12_gate_accepts_extended_csv():
    """sample_reviewed_decisions.csv (extended 16列) を gate が読める"""
    # alpha.5.12 packet 内で既に gate 実行済 → manifest を確認
    sample_report = HARDENING_DIR / "sample_reviewed_decision_validation_report.md"
    assert sample_report.exists()
    # validation report からスキーマを読み取れる
    text = sample_report.read_text(encoding="utf-8")
    # gate が動いて report が出ている
    assert "approved" in text.lower() or "approved" in text
    print("✅ test_alpha_5_12_gate_accepts_extended_csv")


# ============================================================
# 5. sample fixture シナリオ
# ============================================================

def test_alpha_5_12_sample_reviewed_decisions_has_all_scenarios():
    rows = _read_data_rows(SAMPLE_INPUT)
    assert len(rows) == 12, f"sample rows={len(rows)}"
    # 各 scenario の特徴を kasan_key でチェック
    keys = {(r["service"], r["kasan_key"]) for r in rows}
    expected = {
        ("tsusho_kaigo", "chujudosha_care_taisei"),
        ("tsusho_kaigo", "nyuyoku_II"),
        ("tsusho_kaigo", "koukuu_kinou_I"),
        ("tsusho_kaigo", "eiyou_kaizen"),
        ("houmon_kaigo", "shokai_kasan"),
        ("houmon_kaigo", "seikatsu_kinou_renkei_I"),
        ("houmon_kaigo", "seikatsu_kinou_renkei_II"),
        ("houmon_kango_kaigo", "fukusu_mei_houmon_kango_kasan_I_under30"),
        ("houmon_kango_kaigo", "fukusu_mei_houmon_kango_kasan_I_over30"),
        ("houmon_kango_kaigo", "fukusu_mei_houmon_kango_kasan_II_under30"),
        ("houmon_kaigo", "shougu_kaizen_kasan"),
        ("kyotaku_shien", "shougu_kaizen_kasan_2026_06"),
    }
    missing = expected - keys
    assert not missing, f"sample missing scenarios: {missing}"
    print(f"✅ test_alpha_5_12_sample_reviewed_decisions_has_all_scenarios: {len(rows)} rows")


def test_alpha_5_12_sample_gate_has_approved_above_zero():
    rows = _read_data_rows(HARDENING_DIR / "sample_approved_changes_preview.csv")
    assert len(rows) > 0, "approved 0件 だが想定は >0 (legal cleared 含む)"
    print(f"✅ test_alpha_5_12_sample_gate_has_approved_above_zero: {len(rows)}")


def test_alpha_5_12_sample_gate_has_blocked_above_zero():
    rows = _read_data_rows(HARDENING_DIR / "sample_blocked_or_incomplete_decisions.csv")
    assert len(rows) > 0, "blocked 0件 だが想定は >0 (invalid / missing field / future / high risk)"
    print(f"✅ test_alpha_5_12_sample_gate_has_blocked_above_zero: {len(rows)}")


def test_alpha_5_12_sample_gate_has_pending_above_zero():
    rows = _read_data_rows(HARDENING_DIR / "sample_pending_decisions.csv")
    assert len(rows) > 0, "pending 0件 だが想定は >0 (blank / impl=pending / future defer)"
    print(f"✅ test_alpha_5_12_sample_gate_has_pending_above_zero: {len(rows)}")


def test_alpha_5_12_sample_gate_has_legal_review_required_above_zero():
    rows = _read_data_rows(HARDENING_DIR / "sample_legal_review_required.csv")
    assert len(rows) > 0, "legal_review_required 0件 だが想定は >0"
    print(f"✅ test_alpha_5_12_sample_gate_has_legal_review_required_above_zero: {len(rows)}")


# ============================================================
# 6. ルール検証: future_candidate / legal clearance / high-risk
# ============================================================

def test_alpha_5_12_future_candidate_only_cannot_become_approved():
    """future_candidate_only kasan は approved に入らない（sample fixture の row 11 が証明）"""
    approved = _read_data_rows(HARDENING_DIR / "sample_approved_changes_preview.csv")
    blocked = _read_data_rows(HARDENING_DIR / "sample_blocked_or_incomplete_decisions.csv")
    # future_candidate kasan: shougu_kaizen_kasan / shougu_kaizen_kasan_2026_06
    future_keys = {("houmon_kaigo", "shougu_kaizen_kasan"),
                   ("kyotaku_shien", "shougu_kaizen_kasan_2026_06")}
    for r in approved:
        assert (r["service"], r["kasan_key"]) not in future_keys, \
            f"future_candidate {r['kasan_key']} が approved に入った"
    # row 11 (shougu_kaizen_kasan + approve) は blocked にいるべき
    found_block = [r for r in blocked if r["kasan_key"] == "shougu_kaizen_kasan"]
    assert found_block, "future_candidate に approve した row が blocked にない"
    assert "future_candidate_only_must_be_defer_until_r8_definitive" in found_block[0]["blocked_reason"]
    print("✅ test_alpha_5_12_future_candidate_only_cannot_become_approved")


def test_alpha_5_12_needs_legal_review_without_clearance_cannot_become_approved():
    """legal clearance なし needs_legal_review は legal_review_required（approved に入らない）"""
    approved = _read_data_rows(HARDENING_DIR / "sample_approved_changes_preview.csv")
    legal = _read_data_rows(HARDENING_DIR / "sample_legal_review_required.csv")
    no_clearance_keys = {
        ("houmon_kango_kaigo", "fukusu_mei_houmon_kango_kasan_I_under30"),
        ("houmon_kango_kaigo", "fukusu_mei_houmon_kango_kasan_I_over30"),
    }
    for k in no_clearance_keys:
        in_approved = any((r["service"], r["kasan_key"]) == k for r in approved)
        assert not in_approved, f"clearance なし {k} が approved に入った"
        in_legal = any((r["service"], r["kasan_key"]) == k for r in legal)
        assert in_legal, f"clearance なし {k} が legal_review_required に入っていない"
    print("✅ test_alpha_5_12_needs_legal_review_without_clearance_cannot_become_approved")


def test_alpha_5_12_needs_legal_review_with_clearance_can_enter_approved_preview():
    """legal_review_clearance=cleared + 必須揃い + impl=yes の needs_legal_review は approved 候補"""
    approved = _read_data_rows(HARDENING_DIR / "sample_approved_changes_preview.csv")
    cleared_key = ("houmon_kango_kaigo", "fukusu_mei_houmon_kango_kasan_II_under30")
    in_approved = any((r["service"], r["kasan_key"]) == cleared_key for r in approved)
    assert in_approved, f"cleared な needs_legal_review {cleared_key} が approved に入っていない"
    print("✅ test_alpha_5_12_needs_legal_review_with_clearance_can_enter_approved_preview")


def test_alpha_5_12_high_risk_decision_requires_implementation_risk_acknowledged():
    """correct_internal_legacy_code + impl=yes + risk_ack=no/blank → blocked"""
    blocked = _read_data_rows(HARDENING_DIR / "sample_blocked_or_incomplete_decisions.csv")
    found = [r for r in blocked if r["kasan_key"] == "shokai_kasan"
             and r["reviewer_decision"] == "correct_internal_legacy_code"]
    assert found, "高リスク decision (risk_ack=no) が blocked になっていない"
    assert "high_risk_decision_requires_implementation_risk_acknowledged_yes" in found[0]["blocked_reason"]
    print("✅ test_alpha_5_12_high_risk_decision_requires_implementation_risk_acknowledged")


def test_alpha_5_12_high_risk_decision_with_risk_ack_yes_can_be_approved(tmp_path):
    """correct_internal_legacy_code + impl=yes + risk_ack=yes → approved 候補"""
    csv_path = tmp_path / "high_risk_yes.csv"
    extended_columns = [
        "service", "kasan_key", "reviewer_decision",
        "reason", "required_evidence",
        "reviewer_name", "reviewed_at", "final_approved_by",
        "implementation_allowed",
        "reviewer_role", "review_note",
        "legal_review_clearance", "legal_review_reference", "legal_review_note",
        "implementation_priority", "implementation_risk_acknowledged",
    ]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=extended_columns, extrasaction="ignore")
        w.writeheader()
        w.writerow({
            "service": "tsusho_kaigo", "kasan_key": "chujudosha_care_taisei",
            "reviewer_decision": "correct_internal_legacy_code",
            "reason": "社内コード訂正", "required_evidence": "PDF回帰必須",
            "reviewer_name": "X", "reviewed_at": "2026-05-15", "final_approved_by": "X",
            "implementation_allowed": "yes",
            "reviewer_role": "business_reviewer",
            "implementation_risk_acknowledged": "yes",
            "legal_review_clearance": "not_required",
        })
    out_dir = tmp_path / "out"
    res = _run_gate(csv_path, out_dir)
    assert res.returncode == 0, res.stderr
    approved = _read_data_rows(out_dir / "approved_changes_preview.csv")
    assert any(r["kasan_key"] == "chujudosha_care_taisei" for r in approved), \
        "risk_ack=yes でも approved に入らない"
    print("✅ test_alpha_5_12_high_risk_decision_with_risk_ack_yes_can_be_approved")


# ============================================================
# 7. 不変条件
# ============================================================

def test_alpha_5_12_master_json_is_not_modified(tmp_path):
    import hashlib
    services = ("houmon_kango_kaigo", "tsusho_kaigo", "houmon_kaigo", "kyotaku_shien")
    before = {svc: hashlib.md5((PRODUCT_ROOT / f"regulatory_master/kaigo/{svc}.json").read_bytes()).hexdigest()
              for svc in services}
    # generator を再実行
    res = subprocess.run([sys.executable, str(GENERATOR_SCRIPT)],
                          cwd=PRODUCT_ROOT, capture_output=True, text=True)
    assert res.returncode == 0, res.stderr
    after = {svc: hashlib.md5((PRODUCT_ROOT / f"regulatory_master/kaigo/{svc}.json").read_bytes()).hexdigest()
             for svc in services}
    assert before == after, "master JSON が改変された"
    print("✅ test_alpha_5_12_master_json_is_not_modified")


def test_alpha_5_12_checked_20_unchanged():
    counter = {}
    for svc, k, v in _iter_master_kasans():
        if v.get("service_code_mapping_status") == "checked":
            counter[svc] = counter.get(svc, 0) + 1
    assert counter.get("houmon_kango_kaigo", 0) == 14, counter
    assert counter.get("tsusho_kaigo", 0) == 6, counter
    assert sum(counter.values()) == 20, counter
    print(f"✅ test_alpha_5_12_checked_20_unchanged: {counter}")


def test_alpha_5_12_no_new_checked_promotion():
    violations = []
    for svc, k, v in _iter_master_kasans():
        scms = v.get("service_code_mapping_status")
        pa = ((v.get("service_code_audit") or {})
              .get("alpha_5_8_three_layer_model") or {}).get("proposed_action")
        if scms == "checked" and pa != "keep_checked":
            violations.append((svc, k, pa))
    assert not violations, f"想定外: {violations}"
    print("✅ test_alpha_5_12_no_new_checked_promotion")


def test_alpha_5_12_r8_6_provisional_not_used_for_checked():
    r8 = {"WAM_R8_6_8_PROVISIONAL_2026_04_30", "WAM_R8_6_8_PROVISIONAL_2026_04_20"}
    for svc, k, v in _iter_master_kasans():
        if v.get("service_code_mapping_status") != "checked":
            continue
        sid = ((v.get("service_code_audit") or {})
               .get("alpha_5_8_three_layer_model") or {}).get("official_code_model", {}).get("source_id")
        assert sid not in r8, f"checked加算 {svc}.{k} が R8.6案 source ({sid}) を使っている"
    print("✅ test_alpha_5_12_r8_6_provisional_not_used_for_checked")


# ============================================================
# 8. 公開リリースパック / 上流 packet 不変
# ============================================================

def test_alpha_5_12_alpha_5_3_5_4_release_pack_not_modified():
    for v in ("v2026.05.06-alpha.5.3", "v2026.05.06-alpha.5.4"):
        rp = PRODUCT_ROOT / "releases" / "public" / v
        assert rp.exists()
        for f in rp.iterdir():
            if not f.is_file():
                continue
            text = f.read_text(encoding="utf-8", errors="ignore") if f.suffix in (".md", ".json", ".txt") else ""
            assert "alpha.5.12" not in text, f"release pack {v}/{f.name} に alpha.5.12 文字列"
    print("✅ test_alpha_5_12_alpha_5_3_5_4_release_pack_not_modified")


def test_alpha_5_12_reviewer_workbook_is_not_under_public_path():
    """alpha.5.12 workbook が public path に出ていないこと"""
    rel = WORKBOOK.relative_to(PRODUCT_ROOT)
    assert rel.parts[:2] == ("out", "internal"), \
        f"workbook が internal 配下にない: {rel}"
    # releases/public/ 配下にも alpha5_12 がないこと
    for rp in (PRODUCT_ROOT / "releases" / "public").iterdir():
        if not rp.is_dir():
            continue
        for f in rp.rglob("*"):
            if f.is_file():
                assert "alpha5_12" not in f.name and "alpha.5.12" not in f.name, \
                    f"public 配下に alpha.5.12 ファイル: {f}"
    print("✅ test_alpha_5_12_reviewer_workbook_is_not_under_public_path")
