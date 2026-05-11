"""alpha.5.11 reviewer handoff workbook tests.

generate_alpha5_11_reviewer_workbook.py と export_alpha5_11_workbook_decisions.py の
動作と不変条件を保護する。
"""
from __future__ import annotations

import csv
import io
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2].parent
PRODUCT_ROOT = ROOT / "products" / "kasan-manager"
HANDOFF_DIR = PRODUCT_ROOT / "out" / "internal" / "alpha5_11_reviewer_handoff_workbook"
WORKBOOK = HANDOFF_DIR / "alpha5_11_reviewer_decision_workbook.xlsx"
EXPORT_TEMPLATE = HANDOFF_DIR / "reviewer_decision_export_template.csv"
EXPORT_CSV = HANDOFF_DIR / "reviewer_decision_export.csv"

ALPHA_5_9_PACKET = PRODUCT_ROOT / "out" / "internal" / "alpha5_9_master_review_packet"
ALPHA_5_10_GATE = PRODUCT_ROOT / "out" / "internal" / "alpha5_10_reviewer_decision_gate"
ALPHA_5_10_COMMIT = "c9cf05cf5bba29f91231837c34aa42c91153cb27"

GENERATOR_SCRIPT = PRODUCT_ROOT / "scripts" / "generate_alpha5_11_reviewer_workbook.py"
EXPORT_SCRIPT = PRODUCT_ROOT / "scripts" / "export_alpha5_11_workbook_decisions.py"
GATE_SCRIPT = PRODUCT_ROOT / "scripts" / "generate_alpha5_10_reviewer_decision_gate.py"


# ============================================================
# Helpers
# ============================================================

def _read_data_rows(csv_path: Path) -> list[dict]:
    if not csv_path.exists():
        return []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [r for r in reader if r.get("service") and not r.get("service", "").startswith("#")]


def _read_manifest(out_dir: Path, name: str) -> dict:
    return json.loads((out_dir / name).read_text(encoding="utf-8"))


def _iter_master_kasans():
    services = ("houmon_kango_kaigo", "tsusho_kaigo", "houmon_kaigo", "kyotaku_shien")
    for svc in services:
        path = PRODUCT_ROOT / f"regulatory_master/kaigo/{svc}.json"
        d = json.loads(path.read_text(encoding="utf-8"))
        for k, v in (d.get("kasans") or {}).items():
            yield svc, k, v


# ============================================================
# 1. パケット構造
# ============================================================

def test_alpha_5_11_handoff_directory_exists():
    assert HANDOFF_DIR.exists() and HANDOFF_DIR.is_dir()
    expected = {
        "README.md",
        "reviewer_handoff_guide.md",
        "alpha5_11_reviewer_decision_workbook.xlsx",
        "reviewer_decision_export_template.csv",
        "workbook_export_instructions.md",
        "alpha5_11_reviewer_handoff_manifest.json",
    }
    found = {p.name for p in HANDOFF_DIR.iterdir() if p.is_file()}
    missing = expected - found
    assert not missing, f"missing: {missing}"
    print("✅ test_alpha_5_11_handoff_directory_exists")


def test_alpha_5_11_manifest_json_is_valid():
    m = _read_manifest(HANDOFF_DIR, "alpha5_11_reviewer_handoff_manifest.json")
    for key in (
        "version", "base_commit", "input_packet_version", "input_gate_version",
        "total_review_rows", "needs_master_review_count", "needs_legal_review_count",
        "divergent_count", "future_candidate_count",
        "workbook_created", "export_script_created",
        "public_release", "checked_promotion", "master_auto_update",
        "r8_provisional_used_for_checked", "release_pack_modified",
    ):
        assert key in m, f"manifest missing: {key}"
    assert m["version"] == "alpha.5.11"
    assert m["workbook_created"] is True
    assert m["export_script_created"] is True
    assert m["public_release"] is False
    assert m["checked_promotion"] is False
    assert m["master_auto_update"] is False
    assert m["r8_provisional_used_for_checked"] is False
    assert m["release_pack_modified"] is False
    print("✅ test_alpha_5_11_manifest_json_is_valid")


def test_alpha_5_11_manifest_base_commit_equals_alpha_5_10():
    m = _read_manifest(HANDOFF_DIR, "alpha5_11_reviewer_handoff_manifest.json")
    assert m["base_commit"] == ALPHA_5_10_COMMIT
    assert m["input_packet_version"] == "alpha.5.9"
    assert m["input_gate_version"] == "alpha.5.10"
    print("✅ test_alpha_5_11_manifest_base_commit_equals_alpha_5_10")


# ============================================================
# 2. workbook 構造
# ============================================================

def test_alpha_5_11_workbook_xlsx_exists():
    assert WORKBOOK.exists() and WORKBOOK.is_file()
    assert WORKBOOK.stat().st_size > 1000
    print(f"✅ test_alpha_5_11_workbook_xlsx_exists: size={WORKBOOK.stat().st_size}")


def test_alpha_5_11_workbook_has_required_sheets():
    wb = load_workbook(WORKBOOK, read_only=True)
    expected = {
        "README", "Decision_Input_All",
        "Needs_Master_Review", "Needs_Legal_Review",
        "Divergent", "Future_Candidate",
        "Valid_Values", "Gate_Instructions",
    }
    found = set(wb.sheetnames)
    missing = expected - found
    assert not missing, f"missing sheets: {missing}"
    print(f"✅ test_alpha_5_11_workbook_has_required_sheets: {sorted(found)}")


def test_alpha_5_11_decision_input_all_has_38_review_rows():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Decision_Input_All"]
    # row 1 = header / row 2 = guidance / rows 3+ = data
    n = 0
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri <= 2:
            continue
        if row[0]:
            n += 1
    assert n == 38, f"Decision_Input_All rows={n}"
    print(f"✅ test_alpha_5_11_decision_input_all_has_38_review_rows")


def test_alpha_5_11_needs_master_review_has_28_rows():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Needs_Master_Review"]
    n = 0
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri <= 2:
            continue
        if row[0]:
            n += 1
    assert n == 28, f"Needs_Master_Review rows={n}"
    print(f"✅ test_alpha_5_11_needs_master_review_has_28_rows")


def test_alpha_5_11_needs_legal_review_has_5_rows():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Needs_Legal_Review"]
    n = 0
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri <= 2:
            continue
        if row[0]:
            n += 1
    assert n == 5, f"Needs_Legal_Review rows={n}"
    print(f"✅ test_alpha_5_11_needs_legal_review_has_5_rows")


def test_alpha_5_11_divergent_has_3_rows():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Divergent"]
    n = 0
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri <= 2:
            continue
        if row[0]:
            n += 1
    assert n == 3, f"Divergent rows={n}"
    print(f"✅ test_alpha_5_11_divergent_has_3_rows")


def test_alpha_5_11_future_candidate_has_2_rows():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Future_Candidate"]
    n = 0
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri <= 2:
            continue
        if row[0]:
            n += 1
    assert n == 2, f"Future_Candidate rows={n}"
    print(f"✅ test_alpha_5_11_future_candidate_has_2_rows")


def test_alpha_5_11_valid_values_includes_8_decisions_and_3_impl():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Valid_Values"]
    expected_decisions = {
        "approve_official_code_addition", "keep_legacy_detection_only",
        "add_receipt_alias", "correct_internal_legacy_code",
        "mark_structural_mismatch", "keep_pattern_based_unverified",
        "escalate_legal_review", "defer_until_r8_definitive",
    }
    expected_impl = {"yes", "no", "pending"}
    found_decisions = set()
    found_impl = set()
    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        cat, val = row[0], row[1]
        if cat == "reviewer_decision":
            found_decisions.add(val)
        elif cat == "implementation_allowed":
            found_impl.add(val)
    missing_d = expected_decisions - found_decisions
    missing_i = expected_impl - found_impl
    assert not missing_d, f"missing decisions: {missing_d}"
    assert not missing_i, f"missing impl: {missing_i}"
    print(f"✅ test_alpha_5_11_valid_values_includes_8_decisions_and_3_impl")


# ============================================================
# 3. export script
# ============================================================

def test_alpha_5_11_export_script_creates_reviewer_decision_export_csv(tmp_path):
    out = tmp_path / "reviewer_decision_export.csv"
    res = subprocess.run(
        [sys.executable, str(EXPORT_SCRIPT),
         "--workbook", str(WORKBOOK), "--output", str(out)],
        cwd=PRODUCT_ROOT, capture_output=True, text=True,
    )
    assert res.returncode == 0, f"export failed: {res.stderr}"
    assert out.exists()
    print(f"✅ test_alpha_5_11_export_script_creates_reviewer_decision_export_csv")


def test_alpha_5_11_export_csv_has_38_rows(tmp_path):
    out = tmp_path / "reviewer_decision_export.csv"
    subprocess.run([sys.executable, str(EXPORT_SCRIPT),
                    "--workbook", str(WORKBOOK), "--output", str(out)],
                    cwd=PRODUCT_ROOT, capture_output=True, text=True, check=True)
    rows = _read_data_rows(out)
    assert len(rows) == 38, f"export CSV rows={len(rows)}"
    print(f"✅ test_alpha_5_11_export_csv_has_38_rows")


def test_alpha_5_11_export_csv_has_required_columns(tmp_path):
    out = tmp_path / "reviewer_decision_export.csv"
    subprocess.run([sys.executable, str(EXPORT_SCRIPT),
                    "--workbook", str(WORKBOOK), "--output", str(out)],
                    cwd=PRODUCT_ROOT, capture_output=True, text=True, check=True)
    with open(out, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
    expected = {
        "service", "kasan_key", "reviewer_decision",
        "reason", "required_evidence",
        "reviewer_name", "reviewed_at", "final_approved_by",
        "implementation_allowed",
    }
    missing = expected - set(header)
    assert not missing, f"export CSV missing: {missing}"
    print(f"✅ test_alpha_5_11_export_csv_has_required_columns")


# ============================================================
# 4. alpha.5.10 gate との連携
# ============================================================

def test_alpha_5_11_alpha_5_10_gate_default_behavior_still_works(tmp_path):
    """default 引数（input/output 未指定）でも従来どおり動く"""
    res = subprocess.run([sys.executable, str(GATE_SCRIPT)],
                          cwd=PRODUCT_ROOT, capture_output=True, text=True)
    assert res.returncode == 0, res.stderr
    # default 出力先が更新されていること
    m = _read_manifest(ALPHA_5_10_GATE, "alpha5_10_reviewer_decision_gate_manifest.json")
    assert m["total_review_rows"] == 38
    assert m["input_packet_version"] == "alpha.5.9"
    print("✅ test_alpha_5_11_alpha_5_10_gate_default_behavior_still_works")


def test_alpha_5_11_alpha_5_10_gate_optional_workbook_input_works(tmp_path):
    """workbook export CSV を gate の入力にできる"""
    export_out = tmp_path / "wb_export.csv"
    subprocess.run([sys.executable, str(EXPORT_SCRIPT),
                    "--workbook", str(WORKBOOK), "--output", str(export_out)],
                    cwd=PRODUCT_ROOT, capture_output=True, text=True, check=True)
    gate_out = tmp_path / "gate_out"
    res = subprocess.run([sys.executable, str(GATE_SCRIPT),
                           "--input", str(export_out), "--output", str(gate_out)],
                          cwd=PRODUCT_ROOT, capture_output=True, text=True)
    assert res.returncode == 0, res.stderr
    m = _read_manifest(gate_out, "alpha5_10_reviewer_decision_gate_manifest.json")
    assert m["total_review_rows"] == 38
    # blank workbook なので 0 approved / 0 blocked / 38 pending
    assert m["approved_count"] == 0
    assert m["pending_count"] == 38
    print(f"✅ test_alpha_5_11_alpha_5_10_gate_optional_workbook_input_works")


# ============================================================
# 5. 不変条件
# ============================================================

def test_alpha_5_11_master_json_is_not_modified(tmp_path):
    import hashlib
    services = ("houmon_kango_kaigo", "tsusho_kaigo", "houmon_kaigo", "kyotaku_shien")
    before = {svc: hashlib.md5((PRODUCT_ROOT / f"regulatory_master/kaigo/{svc}.json").read_bytes()).hexdigest()
              for svc in services}
    # generator + export を再実行
    subprocess.run([sys.executable, str(GENERATOR_SCRIPT)],
                    cwd=PRODUCT_ROOT, capture_output=True, text=True, check=True)
    out = tmp_path / "x.csv"
    subprocess.run([sys.executable, str(EXPORT_SCRIPT),
                    "--workbook", str(WORKBOOK), "--output", str(out)],
                    cwd=PRODUCT_ROOT, capture_output=True, text=True, check=True)
    after = {svc: hashlib.md5((PRODUCT_ROOT / f"regulatory_master/kaigo/{svc}.json").read_bytes()).hexdigest()
             for svc in services}
    assert before == after, "master JSON が改変された"
    print("✅ test_alpha_5_11_master_json_is_not_modified")


def test_alpha_5_11_checked_20_unchanged():
    counter = {}
    for svc, k, v in _iter_master_kasans():
        if v.get("service_code_mapping_status") == "checked":
            counter[svc] = counter.get(svc, 0) + 1
    assert counter.get("houmon_kango_kaigo", 0) == 14, counter
    assert counter.get("tsusho_kaigo", 0) == 6, counter
    assert sum(counter.values()) == 20, counter
    print(f"✅ test_alpha_5_11_checked_20_unchanged: {counter}")


def test_alpha_5_11_no_new_checked_promotion():
    violations = []
    for svc, k, v in _iter_master_kasans():
        scms = v.get("service_code_mapping_status")
        pa = ((v.get("service_code_audit") or {})
              .get("alpha_5_8_three_layer_model") or {}).get("proposed_action")
        if scms == "checked" and pa != "keep_checked":
            violations.append((svc, k, pa))
    assert not violations, f"想定外: {violations}"
    print("✅ test_alpha_5_11_no_new_checked_promotion")


def test_alpha_5_11_r8_6_provisional_not_used_for_checked():
    r8 = {"WAM_R8_6_8_PROVISIONAL_2026_04_30", "WAM_R8_6_8_PROVISIONAL_2026_04_20"}
    for svc, k, v in _iter_master_kasans():
        if v.get("service_code_mapping_status") != "checked":
            continue
        sid = ((v.get("service_code_audit") or {})
               .get("alpha_5_8_three_layer_model") or {}).get("official_code_model", {}).get("source_id")
        assert sid not in r8, f"checked加算 {svc}.{k} が R8.6案 source ({sid}) を使っている"
    print("✅ test_alpha_5_11_r8_6_provisional_not_used_for_checked")


# ============================================================
# 6. 公開リリースパック / 上流 packet 不変
# ============================================================

def test_alpha_5_11_alpha_5_3_5_4_release_pack_not_modified():
    for v in ("v2026.05.06-alpha.5.3", "v2026.05.06-alpha.5.4"):
        rp = PRODUCT_ROOT / "releases" / "public" / v
        assert rp.exists()
        for f in rp.iterdir():
            if not f.is_file():
                continue
            text = f.read_text(encoding="utf-8", errors="ignore") if f.suffix in (".md", ".json", ".txt") else ""
            assert "alpha.5.11" not in text, f"release pack {v}/{f.name} に alpha.5.11 文字列"
    print("✅ test_alpha_5_11_alpha_5_3_5_4_release_pack_not_modified")


def test_alpha_5_11_alpha_5_9_packet_not_destroyed():
    expected = {
        "README.md", "master_review_summary.md",
        "needs_master_review_matrix.csv", "needs_legal_review_matrix.csv",
        "divergent_mapping_review.md", "future_candidate_review.md",
        "reviewer_decision_template.csv", "alpha5_9_master_review_packet_manifest.json",
    }
    found = {p.name for p in ALPHA_5_9_PACKET.iterdir() if p.is_file()}
    missing = expected - found
    assert not missing, f"alpha.5.9 packet 破壊: {missing}"
    print("✅ test_alpha_5_11_alpha_5_9_packet_not_destroyed")


def test_alpha_5_11_alpha_5_10_gate_not_destroyed():
    expected = {
        "README.md", "decision_validation_report.md",
        "approved_changes_preview.csv", "approved_changes_preview.json",
        "blocked_or_incomplete_decisions.csv", "pending_decisions.csv",
        "legal_review_required.csv", "alpha5_10_reviewer_decision_gate_manifest.json",
    }
    found = {p.name for p in ALPHA_5_10_GATE.iterdir() if p.is_file()}
    missing = expected - found
    assert not missing, f"alpha.5.10 gate 破壊: {missing}"
    print("✅ test_alpha_5_11_alpha_5_10_gate_not_destroyed")


# ============================================================
# 7. safe expressions
# ============================================================

def test_alpha_5_11_handoff_safe_expressions():
    forbidden = [
        "算定可否を保証", "算定を保証", "算定可能と保証",
        "公式コード完全照合済", "完全照合済",
        "R8対応済", "R8.6対応済", "R8.6.1対応済",
    ]
    for f in HANDOFF_DIR.iterdir():
        if not f.is_file():
            continue
        if f.suffix in (".md", ".csv", ".json"):
            text = f.read_text(encoding="utf-8", errors="ignore")
            for w in forbidden:
                assert w not in text, f"{f.name} に禁止語: {w}"
        elif f.suffix == ".xlsx":
            wb = load_workbook(f, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str):
                            for w in forbidden:
                                assert w not in cell, f"workbook[{sn}] に禁止語: {w}"
    print("✅ test_alpha_5_11_handoff_safe_expressions")
