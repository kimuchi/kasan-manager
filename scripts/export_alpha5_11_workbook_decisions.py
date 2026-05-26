"""alpha.5.11 reviewer workbook → CSV export.

reviewer が記入した alpha5_11_reviewer_decision_workbook.xlsx の
Decision_Input_All シートから 38 行を抽出し、alpha.5.10 gate と互換の CSV
形式で reviewer_decision_export.csv を出力する。

方針:
- alpha.5.9 の reviewer_decision_template.csv は **絶対に上書きしない**
- 出力 CSV は alpha.5.10 gate の --input にそのまま渡せる形式
- master JSON は **読み書き共に触らない**
"""
from __future__ import annotations

import argparse
import csv
import io
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = (
    ROOT / "out" / "internal" / "alpha5_11_reviewer_handoff_workbook"
    / "alpha5_11_reviewer_decision_workbook.xlsx"
)
DEFAULT_OUTPUT = (
    ROOT / "out" / "internal" / "alpha5_11_reviewer_handoff_workbook"
    / "reviewer_decision_export.csv"
)

# alpha.5.10 gate と互換の legacy 9列構造
LEGACY_EXPORT_COLUMNS = [
    "service", "kasan_key", "reviewer_decision",
    "reason", "required_evidence",
    "reviewer_name", "reviewed_at", "final_approved_by",
    "implementation_allowed",
]

# alpha.5.12 で導入: 拡張 16列構造（reviewer 備考欄 + legal_review_clearance + risk_ack 等）
EXTENDED_EXPORT_COLUMNS = LEGACY_EXPORT_COLUMNS + [
    "reviewer_role",
    "review_note",
    "legal_review_clearance",
    "legal_review_reference",
    "legal_review_note",
    "implementation_priority",
    "implementation_risk_acknowledged",
]

# 後方互換: alpha.5.11 既存利用者向けに、 default は legacy のまま
EXPORT_COLUMNS = LEGACY_EXPORT_COLUMNS

# Decision_Input_All の列順 (generator と一致)
# alpha.5.11 workbook では 14列 / alpha.5.12 workbook では 21列。
# どちらでも読めるように、ヘッダから動的に列を検出する。
INPUT_SHEET_NAME = "Decision_Input_All"
INPUT_COLUMNS_IN_SHEET_LEGACY = [
    "service", "kasan_key", "kasan_display_name",
    "bucket", "current_overall_mapping_status", "proposed_action",
    "reviewer_decision", "reason", "required_evidence",
    "reviewer_name", "reviewed_at", "final_approved_by",
    "implementation_allowed", "input_status_hint",
]
INPUT_COLUMNS_IN_SHEET_EXTENDED = INPUT_COLUMNS_IN_SHEET_LEGACY + [
    "reviewer_role",
    "review_note",
    "legal_review_clearance",
    "legal_review_reference",
    "legal_review_note",
    "implementation_priority",
    "implementation_risk_acknowledged",
]
# 互換のため両方を許容
INPUT_COLUMNS_IN_SHEET = INPUT_COLUMNS_IN_SHEET_EXTENDED


def export(workbook_path: Path, output_path: Path, force_extended: bool | None = None) -> int:
    """ワークブックの Decision_Input_All シートを CSV にエクスポートする。

    alpha.5.12 拡張: ワークブックに extended 列（reviewer_role / review_note /
    legal_review_clearance / legal_review_reference / legal_review_note /
    implementation_priority / implementation_risk_acknowledged）が含まれる場合は
    それらも CSV に出力する。alpha.5.11 既存ワークブックは 9列のみ出力（後方互換）。

    force_extended:
      - None (default): ワークブック header に拡張列がある場合は extended、なければ legacy
      - True: 必ず extended 16列で出力（拡張列がない場合は空欄で埋める）
      - False: 必ず legacy 9列で出力（拡張列があっても無視）
    """
    if not workbook_path.exists():
        raise SystemExit(f"workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if INPUT_SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"sheet '{INPUT_SHEET_NAME}' not found in {workbook_path}")
    ws = wb[INPUT_SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    # 1行目: ヘッダ / 2行目: ガイダンス merged行 / 3行目以降: データ
    header = next(rows_iter, None)
    _ = next(rows_iter, None)  # ガイダンス行は捨てる
    if header is None:
        raise SystemExit("workbook header missing")

    # 列名→インデックス（拡張列も含めて取得を試みる）
    col_idx = {}
    for i, name in enumerate(header):
        if name in INPUT_COLUMNS_IN_SHEET_EXTENDED:
            col_idx[name] = i

    # 拡張列の存在判定
    has_extended_in_workbook = any(
        c in col_idx for c in (
            "reviewer_role", "review_note",
            "legal_review_clearance", "legal_review_reference", "legal_review_note",
            "implementation_priority", "implementation_risk_acknowledged",
        )
    )
    if force_extended is None:
        use_extended = has_extended_in_workbook
    else:
        use_extended = bool(force_extended)
    columns = EXTENDED_EXPORT_COLUMNS if use_extended else LEGACY_EXPORT_COLUMNS

    def _get(row, name):
        if name not in col_idx:
            return ""
        v = row[col_idx[name]]
        return (v or "").strip() if v is not None else ""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows_iter:
            if row is None:
                continue
            svc = _get(row, "service")
            if not svc:
                continue
            out_row = {col: _get(row, col) for col in columns}
            writer.writerow(out_row)
            n += 1
        # 末尾に注記
        f.write("\n")
        f.write(f"# alpha.5.11/5.12 reviewer workbook export\n")
        f.write(f"# source: {workbook_path.name}\n")
        f.write(f"# schema: {'extended_16_column' if use_extended else 'legacy_9_column'}\n")
        f.write("# alpha.5.9 の reviewer_decision_template.csv は上書きしない\n")
    return n


def main(argv=None):
    parser = argparse.ArgumentParser(description="alpha.5.11/5.12 workbook → CSV exporter")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--schema", choices=["auto", "legacy", "extended"], default="auto",
                        help="出力 CSV のスキーマ: auto (default・workbook header から判定) / legacy (9列) / extended (16列)")
    args = parser.parse_args(argv)

    wb_path = Path(args.workbook)
    out_path = Path(args.output)
    force_extended = None
    if args.schema == "legacy":
        force_extended = False
    elif args.schema == "extended":
        force_extended = True

    n = export(wb_path, out_path, force_extended=force_extended)
    print(f"alpha.5.11/5.12 export: wrote {n} rows to {out_path}")
    print(f"  source: {wb_path}")
    print(f"  schema: {args.schema}")
    print(f"  next : python scripts/generate_alpha5_10_reviewer_decision_gate.py "
          f"--input {out_path} --output out/internal/alpha5_10_reviewer_decision_gate_from_workbook/")


if __name__ == "__main__":
    main()
