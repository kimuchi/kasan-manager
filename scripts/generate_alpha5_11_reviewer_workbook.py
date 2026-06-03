"""alpha.5.11 reviewer handoff workbook generator.

alpha.5.9 master review packet と alpha.5.10 reviewer decision gate を前提に、
業務担当・法令確認者・最終承認者が **Excelで判断入力できる** 内部レビュー用
ワークブック (.xlsx) を生成する。

方針:
- master JSON は **絶対に修正しない**（読み取り専用）
- このワークブックは **内部レビュー用**。public release pack には含めない
- 入力規則（プルダウン・色分け）で reviewer ミスを抑制
- 算定可否保証・公式コード完全照合済み・R8 改定対応完了 表現は禁止
- 入力済 reviewer ファイルは public に出さない（reviewer が記入後はローカル運用）
"""
from __future__ import annotations

import csv
import io
import json
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]

# Versions / commits
PACKET_VERSION = "alpha.5.11"
INPUT_PACKET_VERSION = "alpha.5.9"
INPUT_GATE_VERSION = "alpha.5.10"
ALPHA_5_10_BASE_COMMIT = "c9cf05cf5bba29f91231837c34aa42c91153cb27"
GENERATED_AT = "2026-05-10"

# Paths
ALPHA_5_9_PACKET_DIR = ROOT / "out" / "internal" / "alpha5_9_master_review_packet"
ALPHA_5_10_GATE_DIR = ROOT / "out" / "internal" / "alpha5_10_reviewer_decision_gate"
OUT_DIR = ROOT / "out" / "internal" / "alpha5_11_reviewer_handoff_workbook"
WORKBOOK_PATH = OUT_DIR / "alpha5_11_reviewer_decision_workbook.xlsx"
EXPORT_TEMPLATE_CSV = OUT_DIR / "reviewer_decision_export_template.csv"

# Decision sets (= alpha.5.10 と整合)
VALID_DECISIONS = [
    "approve_official_code_addition",
    "keep_legacy_detection_only",
    "add_receipt_alias",
    "correct_internal_legacy_code",
    "mark_structural_mismatch",
    "keep_pattern_based_unverified",
    "escalate_legal_review",
    "defer_until_r8_definitive",
]
MODIFYING_DECISIONS = {
    "approve_official_code_addition",
    "add_receipt_alias",
    "correct_internal_legacy_code",
}
NON_MODIFYING_DECISIONS = {
    "keep_legacy_detection_only",
    "keep_pattern_based_unverified",
    "mark_structural_mismatch",
    "escalate_legal_review",
    "defer_until_r8_definitive",
}
HIGH_RISK_DECISIONS = {"correct_internal_legacy_code"}
VALID_IMPL_ALLOWED = ["yes", "no", "pending"]

# Decision input columns (= reviewer_decision_template.csv と同一)
DECISION_INPUT_COLUMNS = [
    "service", "kasan_key", "reviewer_decision",
    "reason", "required_evidence",
    "reviewer_name", "reviewed_at", "final_approved_by",
    "implementation_allowed",
]
REQUIRED_FIELDS_WHEN_YES = (
    "reviewer_decision", "reason", "required_evidence",
    "reviewer_name", "reviewed_at", "final_approved_by",
)


# ============================================================
# Style helpers
# ============================================================

def _border():
    side = Side(style="thin", color="888888")
    return Border(left=side, right=side, top=side, bottom=side)

HEADER_FILL = PatternFill("solid", fgColor="305496")  # 紺
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")  # 黄系
WARN_FILL = PatternFill("solid", fgColor="F8CBAD")  # オレンジ系
HIGH_RISK_FILL = PatternFill("solid", fgColor="F4B084")  # 濃オレンジ
NEEDS_LEGAL_FILL = PatternFill("solid", fgColor="DDEBF7")  # 薄水色
FUTURE_CANDIDATE_FILL = PatternFill("solid", fgColor="E2EFDA")  # 薄緑
DIVERGENT_FILL = PatternFill("solid", fgColor="FCE4D6")  # サーモン
REQUIRED_FIELD_FILL = PatternFill("solid", fgColor="FFE699")  # 薄黄
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

CENTER_BOLD = Font(bold=True)


def _set_header(ws, row_idx, columns):
    for ci, col_name in enumerate(columns, start=1):
        c = ws.cell(row=row_idx, column=ci, value=col_name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()


def _autosize(ws, max_width=80, min_width=8):
    for col_cells in ws.columns:
        first = col_cells[0]
        col_letter = get_column_letter(first.column)
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            for line in str(v).split("\n"):
                w = sum(2 if ord(c) > 127 else 1 for c in line)  # CJK 倍幅
                if w > max_len:
                    max_len = w
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


# ============================================================
# Loaders (read-only)
# ============================================================

def _load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [r for r in reader if r.get("service") and not r.get("service", "").startswith("#")]


def load_alpha_5_9_inputs() -> dict:
    return {
        "decision_template": _load_csv(ALPHA_5_9_PACKET_DIR / "reviewer_decision_template.csv"),
        "needs_master_review": _load_csv(ALPHA_5_9_PACKET_DIR / "needs_master_review_matrix.csv"),
        "needs_legal_review": _load_csv(ALPHA_5_9_PACKET_DIR / "needs_legal_review_matrix.csv"),
    }


def load_master_kasans() -> dict:
    services = ("houmon_kango_kaigo", "tsusho_kaigo", "houmon_kaigo", "kyotaku_shien")
    out = {}
    for svc in services:
        path = ROOT / "regulatory_master" / "kaigo" / f"{svc}.json"
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        for k, v in (d.get("kasans") or {}).items():
            out[(svc, k)] = v
    return out


# ============================================================
# Sheet builders
# ============================================================

def build_readme_sheet(wb):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("alpha.5.11 Reviewer Decision Workbook", CENTER_BOLD),
        ("", None),
        (f"version: {PACKET_VERSION}", None),
        (f"base_commit: {ALPHA_5_10_BASE_COMMIT} (alpha.5.10)", None),
        (f"input_packet_version: {INPUT_PACKET_VERSION}", None),
        (f"input_gate_version: {INPUT_GATE_VERSION}", None),
        (f"generated_at: {GENERATED_AT}", None),
        ("", None),
        ("【目的】", CENTER_BOLD),
        ("alpha.5.9 で生成した reviewer_decision_template.csv に、業務担当・法令確認者・最終承認者が", None),
        ("Excel で判断入力するための内部ワークブック。CSV だけでは入力ミスが起きやすいため、", None),
        ("プルダウン・色分け・担当者別シートを用意した。", None),
        ("", None),
        ("【public 資料ではありません】", CENTER_BOLD),
        ("- このブックは out/internal 配下の内部レビュー資料であり、public release pack には含めない", None),
        ("- reviewer が入力した実判断ファイルも public に出さない", None),
        ("- alpha.5.3 / alpha.5.4 release pack には影響しない", None),
        ("", None),
        ("【入力しても自動で master 修正されません】", CENTER_BOLD),
        ("- このブックへの入力は master JSON を改変しません", None),
        ("- export script で CSV 化したあと、alpha.5.10 gate で検証してから次フェーズに進みます", None),
        ("- 新規 checked 昇格・公式コードへの一括置換は alpha.5.11 では一切行いません", None),
        ("- R8.6 案資料は checked 昇格に使いません", None),
        ("", None),
        ("【次にやること】", CENTER_BOLD),
        ("1. Decision_Input_All シートで 38 件の各行に判断を入力", None),
        ("2. 必要に応じて Needs_Master_Review / Needs_Legal_Review / Divergent / Future_Candidate を参照", None),
        ("3. 入力後にこのファイルを保存", None),
        ("4. ターミナルで `python scripts/export_alpha5_11_workbook_decisions.py` を実行", None),
        ("5. 出力された CSV を alpha.5.10 gate で再検証", None),
        ("6. blocked / pending / legal_review_required を解消するまで 1〜5 を繰り返す", None),
        ("7. approved 行が確定したら、開発担当が alpha.5.12+ で master JSON 段階反映を別 PR で実施", None),
        ("", None),
        ("【シート一覧】", CENTER_BOLD),
        ("- README                : 本ファイル", None),
        ("- Decision_Input_All    : 入力シート（38件・プルダウン・色分けあり）", None),
        ("- Needs_Master_Review   : 業務担当向け参照（28件）", None),
        ("- Needs_Legal_Review    : 法令確認者向け参照（5件）", None),
        ("- Divergent             : proposed_action と overall が divergent な3件", None),
        ("- Future_Candidate      : R8.6 確定版待ち2件（必ず defer）", None),
        ("- Valid_Values          : reviewer_decision / implementation_allowed の選択肢一覧", None),
        ("- Gate_Instructions     : Excel入力後の export → gate 再実行手順", None),
    ]
    for r_idx, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=r_idx, column=1, value=text)
        if font:
            c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100
    return ws


def build_valid_values_sheet(wb):
    ws = wb.create_sheet("Valid_Values")
    columns = ["category", "value", "is_modifying", "is_high_risk", "note"]
    _set_header(ws, 1, columns)
    rows = []
    for d in VALID_DECISIONS:
        is_mod = d in MODIFYING_DECISIONS
        is_hr = d in HIGH_RISK_DECISIONS
        note = {
            "approve_official_code_addition": "公式コードを追加する。低リスク。",
            "keep_legacy_detection_only": "社内 legacy code 運用を維持。master 修正なし。",
            "add_receipt_alias": "公式コードを社内コードの alias として登録。中リスク。",
            "correct_internal_legacy_code": "社内 service_codes を公式コードに置換。**高リスク**（PDF検出回帰必須）。",
            "mark_structural_mismatch": "structural_mismatch として明示。法令解釈待ち。",
            "keep_pattern_based_unverified": "パターン検出のまま継続。master 修正なし。",
            "escalate_legal_review": "法令確認者にエスカレーション。approved にならない。",
            "defer_until_r8_definitive": "R8.6.1 確定版が出るまで保留。",
        }.get(d, "")
        rows.append(("reviewer_decision", d, "yes" if is_mod else "no",
                     "yes" if is_hr else "no", note))
    for v in VALID_IMPL_ALLOWED:
        note = {
            "yes": "実装してよい。**6 必須フィールド全揃い**が前提（reviewer_decision/reason/required_evidence/reviewer_name/reviewed_at/final_approved_by）",
            "no": "実装しない。決定のみ記録。",
            "pending": "保留。決定確定前。",
        }[v]
        rows.append(("implementation_allowed", v, "", "", note))
    rows.append(("reviewer_role", "業務担当", "", "", "needs_master_review / divergent の判断"))
    rows.append(("reviewer_role", "法令確認者", "", "", "needs_legal_review の judgment、escalate_legal_review の解析"))
    rows.append(("reviewer_role", "最終承認者", "", "", "approved_changes_preview の最終ハンコ。implementation_allowed=yes の final_approved_by"))
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            if r_idx <= 1 + len(VALID_DECISIONS) and val in HIGH_RISK_DECISIONS:
                c.fill = HIGH_RISK_FILL
    _autosize(ws)
    return ws


def build_decision_input_sheet(wb, decision_template_rows, master_data, alpha_5_9_legal_keys,
                                alpha_5_9_master_keys, divergent_keys, future_candidate_keys):
    ws = wb.create_sheet("Decision_Input_All")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "bucket", "current_overall_mapping_status", "proposed_action",
        "reviewer_decision", "reason", "required_evidence",
        "reviewer_name", "reviewed_at", "final_approved_by",
        "implementation_allowed", "input_status_hint",
    ]
    _set_header(ws, 1, columns)

    # 1行目に注意書き行を入れる（行2 = 注意書き、データは 行3〜）
    note_row = (
        "（注）implementation_allowed=yes の場合は reviewer_decision / reason / required_evidence / reviewer_name / reviewed_at / final_approved_by が必須（黄色マークセル）。",
        "needs_legal_review 5件は approved にならず legal_review_required へ。future_candidate_only 2件は defer_until_r8_definitive のみ可。",
    )
    note_cell = ws.cell(row=2, column=1,
                        value="ガイダンス: " + note_row[0] + "\n" + note_row[1])
    note_cell.fill = NOTE_FILL
    note_cell.alignment = WRAP_ALIGN
    note_cell.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    # 3行目以降にデータ
    start_data_row = 3
    for ri, row in enumerate(decision_template_rows, start=start_data_row):
        svc = row.get("service", "")
        k = row.get("kasan_key", "")
        kasan_def = master_data.get((svc, k)) or {}
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        pa = three.get("proposed_action", "")
        ov = kasan_def.get("overall_mapping_status", "")
        # bucket determination
        if (svc, k) in alpha_5_9_legal_keys:
            bucket = "needs_legal_review"
        elif pa == "future_candidate_only":
            bucket = "future_candidate_only"
        elif (svc, k) in divergent_keys:
            bucket = "divergent (keep_pattern_based_unverified ∧ overall=needs_review)"
        elif pa == "needs_master_review":
            bucket = "needs_master_review"
        else:
            bucket = pa or ""
        hint_parts = []
        if bucket == "needs_legal_review":
            hint_parts.append("⚠ legal review 必須・approved にならない")
        if bucket == "future_candidate_only":
            hint_parts.append("⚠ R8.6.1 確定版待ち・defer_until_r8_definitive のみ可")
        if bucket.startswith("divergent"):
            hint_parts.append("⚠ divergent 加算・alpha.5.8.1 で audit_note 化済")
        hint = " / ".join(hint_parts) or "業務担当が判断・記入"

        cells_data = [
            svc, k, kasan_def.get("name", ""),
            bucket, ov, pa,
            row.get("reviewer_decision", ""),
            row.get("reason", ""),
            row.get("required_evidence", ""),
            row.get("reviewer_name", ""),
            row.get("reviewed_at", ""),
            row.get("final_approved_by", ""),
            row.get("implementation_allowed", ""),
            hint,
        ]
        for ci, val in enumerate(cells_data, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()

        # bucket-color
        bucket_cell = ws.cell(row=ri, column=4)
        if bucket == "needs_legal_review":
            for ci in range(1, len(columns) + 1):
                ws.cell(row=ri, column=ci).fill = NEEDS_LEGAL_FILL
        elif bucket == "future_candidate_only":
            for ci in range(1, len(columns) + 1):
                ws.cell(row=ri, column=ci).fill = FUTURE_CANDIDATE_FILL
        elif bucket.startswith("divergent"):
            for ci in range(1, len(columns) + 1):
                ws.cell(row=ri, column=ci).fill = DIVERGENT_FILL

        # required-field 黄色マーキング: 入力欄 (col 7-12) を薄黄でマーク
        # (impl=yes 時に必要であることを視覚的に示す)
        for col_idx in (7, 8, 9, 10, 11, 12):
            c = ws.cell(row=ri, column=col_idx)
            # 既に bucket 色が付いている場合はそれを優先
            if c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = REQUIRED_FIELD_FILL
        # implementation_allowed (col 13) も同様
        impl_cell = ws.cell(row=ri, column=13)
        if impl_cell.fill.fgColor.rgb in (None, "00000000"):
            impl_cell.fill = REQUIRED_FIELD_FILL

    n_data_rows = len(decision_template_rows)
    last_row = start_data_row + n_data_rows - 1

    # データ検証 (プルダウン)
    decision_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(VALID_DECISIONS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="無効な reviewer_decision",
        error="VALID_DECISIONS シートの値から選択してください",
    )
    decision_dv.add(f"G{start_data_row}:G{last_row}")
    ws.add_data_validation(decision_dv)

    impl_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(VALID_IMPL_ALLOWED) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="無効な implementation_allowed",
        error="yes / no / pending のいずれかを選択してください",
    )
    impl_dv.add(f"M{start_data_row}:M{last_row}")
    ws.add_data_validation(impl_dv)

    ws.freeze_panes = f"D{start_data_row}"
    _autosize(ws, max_width=40)
    return ws


def build_needs_master_review_sheet(wb, rows, master_data):
    ws = wb.create_sheet("Needs_Master_Review")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "official_service_code", "official_name", "official_unit", "official_calc_unit",
        "official_source_id",
        "receipt_detection_code", "receipt_detection_pattern", "receipt_detection_status",
        "internal_legacy_code", "internal_legacy_unit",
        "mismatch_type", "mismatch_summary",
        "proposed_review_question", "recommended_next_step",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(row=2, column=1, value="（業務担当向け）公式コードと社内 legacy コードの照合を行ってください。decision を入力するのは Decision_Input_All シートです。")
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    for ri, row in enumerate(rows, start=3):
        cells = [row.get(c, "") for c in columns]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()

    ws.freeze_panes = "D3"
    _autosize(ws, max_width=50)
    return ws


def build_needs_legal_review_sheet(wb, rows):
    ws = wb.create_sheet("Needs_Legal_Review")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "legal_review_reason", "structural_issue_type",
        "official_source_id", "internal_legacy_code", "receipt_detection_pattern",
        "current_overall_mapping_status",
        "why_not_checked", "legal_question", "reference_needed",
        "recommended_next_step",
        "legal_review_clearance_note (optional・備考のみ)",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(row=2, column=1, value="（法令確認者向け）legal_review_clearance フラグは alpha.5.11 では実装しません。備考欄に確認結果を残し、最終決定は Decision_Input_All シートに記入してください。escalate_legal_review にしても approved になりません（legal_review_required に分離）。")
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    for ri, row in enumerate(rows, start=3):
        cells = [
            row.get("service", ""), row.get("kasan_key", ""), row.get("kasan_display_name", ""),
            row.get("legal_review_reason", ""), row.get("structural_issue_type", ""),
            row.get("official_source_id", ""), row.get("internal_legacy_code", ""),
            row.get("receipt_detection_pattern", ""),
            row.get("current_overall_mapping_status", ""),
            row.get("why_not_checked", ""), row.get("legal_question", ""),
            row.get("reference_needed", ""), row.get("recommended_next_step", ""),
            "",  # 法令確認者の備考欄（手書き）
        ]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            c.fill = NEEDS_LEGAL_FILL
        # 備考欄だけ薄黄
        ws.cell(row=ri, column=14).fill = REQUIRED_FIELD_FILL

    ws.freeze_panes = "D3"
    _autosize(ws, max_width=50)
    return ws


def build_divergent_sheet(wb, master_data, divergent_keys):
    ws = wb.create_sheet("Divergent")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "proposed_action", "overall_mapping_status",
        "official_match_type", "official_code", "internal_legacy_code",
        "divergence_reason",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(row=2, column=1, value="（参照のみ）alpha.5.8.1 で audit_note 化済の divergent 3件。proposed_action と overall_mapping_status が分岐している理由を確認してから Decision_Input_All で判断してください。本シートでは判断は入れません。")
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    ri = 3
    for key in sorted(divergent_keys):
        kasan_def = master_data.get(key)
        if not kasan_def:
            continue
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        official = three.get("official_code_model") or {}
        legacy = three.get("internal_legacy_model") or {}
        note_obj = three.get("alpha_5_8_1_proposed_overall_divergence_note") or {}
        cells = [
            key[0], key[1], kasan_def.get("name", ""),
            three.get("proposed_action", ""),
            kasan_def.get("overall_mapping_status", ""),
            official.get("official_match_type", ""),
            official.get("official_service_code") or "",
            legacy.get("internal_legacy_code") or "",
            note_obj.get("reason", ""),
        ]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            c.fill = DIVERGENT_FILL
        ri += 1

    ws.freeze_panes = "D3"
    _autosize(ws, max_width=80)
    return ws


def build_future_candidate_sheet(wb, master_data, future_candidate_keys):
    ws = wb.create_sheet("Future_Candidate")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "proposed_action", "overall_mapping_status",
        "official_source_id", "audit_note", "required_decision",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(
        row=2, column=1,
        value=("（参照のみ）future_candidate_only 2件。R8.6.1 確定版が出るまで "
               "**reviewer_decision は defer_until_r8_definitive のみ**受理します。"
               " R8.6 案資料 (WAM_R8_6_8_PROVISIONAL_2026_04_30) は checked 昇格に使いません。"))
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    ri = 3
    for key in sorted(future_candidate_keys):
        kasan_def = master_data.get(key)
        if not kasan_def:
            continue
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        official = three.get("official_code_model") or {}
        cells = [
            key[0], key[1], kasan_def.get("name", ""),
            three.get("proposed_action", ""),
            kasan_def.get("overall_mapping_status", ""),
            official.get("source_id") or "",
            three.get("audit_note", ""),
            "defer_until_r8_definitive",
        ]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            c.fill = FUTURE_CANDIDATE_FILL
        ri += 1

    ws.freeze_panes = "D3"
    _autosize(ws, max_width=80)
    return ws


def build_gate_instructions_sheet(wb):
    ws = wb.create_sheet("Gate_Instructions")
    blocks = [
        ("Excel 入力後の流れ", CENTER_BOLD),
        ("", None),
        ("1. このブックを保存（.xlsx 形式・上書き）", None),
        ("2. ターミナルでカレントを products/kasan-manager に移動", None),
        ("3. 以下を実行: python scripts/export_alpha5_11_workbook_decisions.py", None),
        ("4. 出力: out/internal/alpha5_11_reviewer_handoff_workbook/reviewer_decision_export.csv", None),
        ("", None),
        ("alpha.5.10 gate を export CSV に対して再実行する方法（optional）:", CENTER_BOLD),
        ("python scripts/generate_alpha5_10_reviewer_decision_gate.py \\", None),
        ("  --input out/internal/alpha5_11_reviewer_handoff_workbook/reviewer_decision_export.csv \\", None),
        ("  --output out/internal/alpha5_10_reviewer_decision_gate_from_workbook/", None),
        ("", None),
        ("blocked_or_incomplete_decisions.csv の見方", CENTER_BOLD),
        ("- blocked_reason: invalid_reviewer_decision / missing_required_fields_when_implementation_allowed_yes / future_candidate_only_must_be_defer_until_r8_definitive 等", None),
        ("- recommended_fix: 修正方法の提案（reviewer_decision の値修正・必須フィールド入力 等）", None),
        ("- 修正後に再度 export → gate 再実行を繰り返す", None),
        ("", None),
        ("pending_decisions.csv の見方", CENTER_BOLD),
        ("- blank_template_row: 入力が必要", None),
        ("- non_modifying_decision_recorded: 決定は記録されたが master 修正対象外", None),
        ("- deferred_until_r8_definitive: R8.6.1 確定版が出るまで保留", None),
        ("- implementation_not_yet_approved: implementation_allowed が yes 以外", None),
        ("", None),
        ("legal_review_required.csv の見方", CENTER_BOLD),
        ("- legal_question を確認し、関係告示・通知を参照", None),
        ("- 解析結果を Decision_Input_All の reviewer_decision / reason / final_approved_by に反映", None),
        ("- escalate_legal_review でも approved にならない（legal_review_required に分離）", None),
    ]
    for r_idx, (text, font) in enumerate(blocks, start=1):
        c = ws.cell(row=r_idx, column=1, value=text)
        if font:
            c.font = font
        c.alignment = WRAP_ALIGN
    ws.column_dimensions["A"].width = 120
    return ws


# ============================================================
# Main
# ============================================================

def write_export_template_csv():
    """空欄の export template CSV を出す（reviewer 入力前の参考形式）。"""
    with open(EXPORT_TEMPLATE_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=DECISION_INPUT_COLUMNS, extrasaction="ignore")
        w.writeheader()
        # alpha.5.9 のテンプレートと同じ 38 行を空欄で書き出す
        rows = _load_csv(ALPHA_5_9_PACKET_DIR / "reviewer_decision_template.csv")
        for r in rows:
            w.writerow({col: r.get(col, "") for col in DECISION_INPUT_COLUMNS})
        f.write("\n")
        f.write("# alpha.5.11 export template (空欄). reviewer は Excel ワークブック側で入力する.\n")
        f.write("# export script: scripts/export_alpha5_11_workbook_decisions.py\n")


def write_readme(out_dir: Path, manifest: dict):
    text = f"""# alpha.5.11 Reviewer Handoff Workbook

**version**: {PACKET_VERSION}
**base_commit**: `{ALPHA_5_10_BASE_COMMIT}` (alpha.5.10 reviewer_decision_gate)
**input_packet_version**: {INPUT_PACKET_VERSION}
**input_gate_version**: {INPUT_GATE_VERSION}
**generated_at**: {GENERATED_AT}

---

## 位置付け

alpha.5.9 master_review_packet と alpha.5.10 reviewer_decision_gate を前提に、
**業務担当・法令確認者・最終承認者が Excel で判断入力できる** 内部レビュー用ワークブック。

- このブックは **out/internal 配下の内部レビュー資料**であり、**public release pack には含めません**
- reviewer が入力した実判断ファイルも public に出しません
- alpha.5.3 / alpha.5.4 release pack には影響しません

## 不変条件（テストで保護）

- ❌ master JSON 自動修正なし（generator は読み取り専用）
- ❌ 新規 checked 昇格なし
- ❌ R8.6 案資料は checked 昇格に使わない
- ❌ public release pack は本 alpha.5.11 で更新しない
- ❌ alpha.5.9 packet / alpha.5.10 gate は破壊しない
- ❌ 過剰な完了感を与える表現は使わない（disclaimer 維持）

## 含まれるファイル

| ファイル | 内容 |
|---|---|
| `README.md` | 本ファイル |
| `reviewer_handoff_guide.md` | 業務担当 / 法令確認者 / 最終承認者向け手順・入力例 |
| `alpha5_11_reviewer_decision_workbook.xlsx` | 8シート構成のレビュー用ワークブック（プルダウン・色分けあり） |
| `reviewer_decision_export_template.csv` | export 後の CSV 形式リファレンス（空欄） |
| `workbook_export_instructions.md` | Excel 保存・export script 実行・gate 再実行の手順 |
| `alpha5_11_reviewer_handoff_manifest.json` | パケットメタデータ |

## ワークブック構成（8 シート）

1. **README** — 本ブックの目的・運用注意
2. **Decision_Input_All** — 38件の入力シート（プルダウン・色分け）
3. **Needs_Master_Review** — 業務担当向け参照（28件）
4. **Needs_Legal_Review** — 法令確認者向け参照（5件）
5. **Divergent** — divergent 3件（参照のみ）
6. **Future_Candidate** — 2件（必ず defer_until_r8_definitive）
7. **Valid_Values** — reviewer_decision / implementation_allowed の選択肢一覧
8. **Gate_Instructions** — Excel 入力後の export → gate 再実行手順

## 次にやること

### reviewer
1. このブックを開いて Decision_Input_All シートで 38 行を埋める
2. 必要に応じて Needs_Master_Review / Needs_Legal_Review / Divergent / Future_Candidate を参照
3. ブックを保存（.xlsx 形式・上書き）
4. ターミナル: `python scripts/export_alpha5_11_workbook_decisions.py`
5. 出力 CSV を alpha.5.10 gate で再検証:
   ```
   python scripts/generate_alpha5_10_reviewer_decision_gate.py \\
     --input out/internal/alpha5_11_reviewer_handoff_workbook/reviewer_decision_export.csv \\
     --output out/internal/alpha5_10_reviewer_decision_gate_from_workbook/
   ```
6. blocked / pending / legal_review_required を解消するまで 1〜5 を繰り返す

### 開発担当
- approved 行が確定したら、**alpha.5.12+** で master JSON 段階反映を別 PR で実施
- **本パッケージは master JSON を改変しません**（reviewer 入力 → export → gate 検証のみ）

## サマリ

| カテゴリ | 件数 |
|---|---:|
| Decision_Input_All の合計 | {manifest['total_review_rows']} |
| Needs_Master_Review | {manifest['needs_master_review_count']} |
| Needs_Legal_Review | {manifest['needs_legal_review_count']} |
| Divergent | {manifest['divergent_count']} |
| Future_Candidate | {manifest['future_candidate_count']} |

## 再生成方法

```
cd products/kasan-manager
python scripts/generate_alpha5_11_reviewer_workbook.py
```
"""
    (out_dir / "README.md").write_text(text, encoding="utf-8")


def write_handoff_guide(out_dir: Path, manifest: dict):
    text = f"""# alpha.5.11 Reviewer Handoff Guide

**version**: {PACKET_VERSION}
**base_commit**: `{ALPHA_5_10_BASE_COMMIT}` (alpha.5.10)
**generated_at**: {GENERATED_AT}

---

## このガイドの目的

alpha.5.11 reviewer ワークブック (`alpha5_11_reviewer_decision_workbook.xlsx`) を使って、
業務担当・法令確認者・最終承認者がそれぞれの役割で **正しく** 判断入力するための手順書です。

## 役割別フロー

### 1. 業務担当向け

#### 担当範囲
- needs_master_review 28件（社内コードと公式コードの不一致）
- divergent 3件（pa と overall の分岐）

#### 手順
1. ワークブックを開いて **Needs_Master_Review** シートで内訳を確認
2. **Decision_Input_All** シートで対象行（bucket = needs_master_review / divergent）を編集
3. reviewer_decision プルダウンから選択:
   - `approve_official_code_addition`（公式コードを master に追加）
   - `add_receipt_alias`（公式コードを社内コードの alias 登録）
   - `correct_internal_legacy_code`（社内コード置換 — **高リスク**）
   - `keep_legacy_detection_only`（現状維持）
   - `mark_structural_mismatch`（構造解釈不要・記録のみ）
4. reason / required_evidence / reviewer_name / reviewed_at を記入
5. final_approved_by は最終承認者の入力枠（自分が最終承認者でない限り空欄でOK）
6. implementation_allowed をプルダウンから選択（yes / no / pending）

#### 入力例（OK）
```
service: tsusho_kaigo
kasan_key: chujudosha_care_taisei
reviewer_decision: add_receipt_alias
reason: 公式コード 155306 と社内コード 156271 の単位は一致するためエイリアス登録で対応
required_evidence: WAM_R7_8_DEFINITIVE_2025_03_28 PDF p131 で確認済
reviewer_name: 業務担当A
reviewed_at: 2026-05-15
final_approved_by: （最終承認者記入）
implementation_allowed: pending
```

#### 入力例（NG）
- `reviewer_decision: approve_official_code_addition` だが reason / required_evidence が空欄
  → `implementation_allowed=yes` にすると blocked
- `reviewer_decision: definitely_invalid_choice` → blocked（プルダウン外の値）
- needs_legal_review 行に `add_receipt_alias` → legal_review_required に分離（approved にならない）

### 2. 法令確認者向け

#### 担当範囲
- needs_legal_review 5件（複数名訪問看護加算 4件 + 長時間訪問看護加算 1件）
- escalate_legal_review が選ばれた他バケットの加算

#### 手順
1. **Needs_Legal_Review** シートで `legal_question` / `reference_needed` / `why_not_checked` を読む
2. 関係告示・通知（介護報酬告示・大臣基準告示・老企第36号 解釈通知 等）を参照
3. 確認結果を **Needs_Legal_Review** シート右端の `legal_review_clearance_note (optional)` 列に書く（備考のみ）
4. **Decision_Input_All** シートで該当行を編集:
   - 法令通知が出ていない場合 → `escalate_legal_review`（approved にならない）
   - 構造的に独立コードがないと判断 → `mark_structural_mismatch`
   - R8.6.1 確定版待ち → `defer_until_r8_definitive`

#### 重要
- alpha.5.11 では `legal_review_clearance` フラグは未実装です
- どんな decision でも、needs_legal_review バケットの行は alpha.5.10 gate で **legal_review_required** に分離されます
- approved にしたい場合は alpha.5.12+ で `legal_review_clearance` フラグを実装してから再評価

### 3. 最終承認者向け

#### 担当範囲
- 業務担当・法令確認者の入力後、`final_approved_by` 列にハンコを入れる
- `implementation_allowed=yes` の最終判断（特に `correct_internal_legacy_code` のような **高リスク** 行）

#### 手順
1. **Decision_Input_All** シートで全38行を確認
2. reviewer_decision / reason / required_evidence / reviewer_name / reviewed_at が揃っているか確認
3. 問題なければ `final_approved_by` に氏名・役職を記入し、`implementation_allowed=yes` に変更
4. `correct_internal_legacy_code` の行は **PDF検出回帰テスト必須** であることを開発担当に伝達
5. 保存後、ターミナルで `python scripts/export_alpha5_11_workbook_decisions.py` を実行
6. CSV 出力を確認後、`scripts/generate_alpha5_10_reviewer_decision_gate.py --input ...` で再ゲート

## R8.6 案資料の扱い（重要）

- WAM_R8_6_8_PROVISIONAL_2026_04_30 は **案資料** で、checked 昇格には使えません
- future_candidate_only 2 件（訪介 shougu_kaizen_kasan / 居宅 shougu_kaizen_kasan_2026_06）は
  必ず `defer_until_r8_definitive` のみ受理します
- 他の decision を入れると alpha.5.10 gate で **blocked** になります
- R8.6.1 確定版が出たら alpha.5.12+ で再評価

## NG decision 例（approved にならない）

| 状況 | 結果 |
|---|---|
| needs_legal_review 行で何でも入力 | legal_review_required へ |
| future_candidate_only 行で `approve_official_code_addition` | blocked（必ず defer） |
| 重複行（service+kasan_key 同じ） | 後発が blocked |
| `implementation_allowed=yes` で必須フィールド欠落 | blocked |
| プルダウン外の値 | blocked |

## サマリ

| カテゴリ | 件数 | 担当 |
|---|---:|---|
| Decision_Input_All 合計 | {manifest['total_review_rows']} | 各役割 |
| needs_master_review | {manifest['needs_master_review_count']} | 業務担当 |
| needs_legal_review | {manifest['needs_legal_review_count']} | 法令確認者 |
| divergent | {manifest['divergent_count']} | 業務担当 |
| future_candidate_only | {manifest['future_candidate_count']} | reviewer 操作不要・defer のみ |
"""
    (out_dir / "reviewer_handoff_guide.md").write_text(text, encoding="utf-8")


def write_export_instructions(out_dir: Path):
    text = f"""# alpha.5.11 Workbook Export Instructions

## Excel 保存時の注意

- **必ず .xlsx 形式で保存**してください（.xls 旧形式 / .csv 単体での保存は禁止）
- 上書き保存で OK（ファイル名は `alpha5_11_reviewer_decision_workbook.xlsx` のまま）
- マクロは使っていないので「ブック有効化」のような選択肢が出た場合は標準のままで保存

## export script の実行方法

```
cd products/kasan-manager
python scripts/export_alpha5_11_workbook_decisions.py
```

オプション:
- `--workbook` : 別の Excel パスを指定（default: `out/internal/alpha5_11_reviewer_handoff_workbook/alpha5_11_reviewer_decision_workbook.xlsx`）
- `--output` : 出力 CSV のパス（default: `out/internal/alpha5_11_reviewer_handoff_workbook/reviewer_decision_export.csv`）

## export 結果の確認

```
out/internal/alpha5_11_reviewer_handoff_workbook/reviewer_decision_export.csv
```

このファイルは alpha.5.10 gate と互換の列構造（service / kasan_key / reviewer_decision / reason / required_evidence / reviewer_name / reviewed_at / final_approved_by / implementation_allowed）です。

## alpha.5.10 gate を再実行する方法（reviewer 入力後）

### default 挙動（alpha.5.9 テンプレート読み込み）
```
python scripts/generate_alpha5_10_reviewer_decision_gate.py
```
→ 入力: `out/internal/alpha5_9_master_review_packet/reviewer_decision_template.csv`
→ 出力: `out/internal/alpha5_10_reviewer_decision_gate/`

### workbook export を入力にする（alpha.5.11 ワークブック由来）
```
python scripts/generate_alpha5_10_reviewer_decision_gate.py \\
  --input out/internal/alpha5_11_reviewer_handoff_workbook/reviewer_decision_export.csv \\
  --output out/internal/alpha5_10_reviewer_decision_gate_from_workbook/
```
→ 出力: `out/internal/alpha5_10_reviewer_decision_gate_from_workbook/`

**default 挙動は変更されません**。reviewer ワークブック由来の検証結果は別ディレクトリに分離されます。

## blocked_or_incomplete_decisions.csv の直し方

| blocked_reason | 直し方 |
|---|---|
| invalid_reviewer_decision | プルダウン外の値が入った。Decision_Input_All の reviewer_decision を Valid_Values シートの値から選び直す |
| invalid_implementation_allowed | yes / no / pending 以外が入った。再選択 |
| missing_required_fields_when_implementation_allowed_yes | implementation_allowed=yes だが必須フィールド欠落。reason / required_evidence / reviewer_name / reviewed_at / final_approved_by を埋める |
| duplicate_service_kasan_key | 同じ kasan_key が複数行ある。1つに統合 |
| future_candidate_only_must_be_defer_until_r8_definitive | future_candidate 行は defer のみ可。reviewer_decision を `defer_until_r8_definitive` に変更 |

## pending_decisions.csv の見方

| pending_reason | 意味 | 次のアクション |
|---|---|---|
| blank_template_row | 未入力 | 業務担当・法令確認者・最終承認者が記入 |
| non_modifying_decision_recorded | 決定は記録されたが master 修正対象外 | 追加作業不要・記録のまま |
| deferred_until_r8_definitive | R8.6.1 確定版待ち | 確定版が出たら alpha.5.12+ で再評価 |
| implementation_not_yet_approved | implementation_allowed が yes 以外 | 最終承認者が yes に変更 |

## legal_review_required.csv の見方

- legal_question / reference_needed が記載されている
- 法令確認者が告示・通知を確認し、Decision_Input_All の該当行を再記入
- ただし alpha.5.11 では `legal_review_clearance` フラグは未実装
- alpha.5.12+ で実装予定（それまでは legal_review_required に滞留）
"""
    (out_dir / "workbook_export_instructions.md").write_text(text, encoding="utf-8")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    inputs = load_alpha_5_9_inputs()
    master_data = load_master_kasans()

    decision_template_rows = inputs["decision_template"]
    needs_master_rows = inputs["needs_master_review"]
    needs_legal_rows = inputs["needs_legal_review"]

    alpha_5_9_legal_keys = {(r["service"], r["kasan_key"]) for r in needs_legal_rows}
    alpha_5_9_master_keys = {(r["service"], r["kasan_key"]) for r in needs_master_rows}

    # divergent / future_candidate keys を master JSON から導出
    divergent_keys = set()
    future_candidate_keys = set()
    for key, kasan_def in master_data.items():
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        pa = three.get("proposed_action")
        ov = kasan_def.get("overall_mapping_status")
        if pa == "future_candidate_only":
            future_candidate_keys.add(key)
        if pa == "keep_pattern_based_unverified" and ov == "needs_review":
            divergent_keys.add(key)

    # Workbook 構築
    wb = Workbook()
    # default sheet を消して new シートを 0 番目から作るので、最初のシート (Sheet) を削除
    default_ws = wb.active
    wb.remove(default_ws)

    build_readme_sheet(wb)
    build_decision_input_sheet(wb, decision_template_rows, master_data,
                                alpha_5_9_legal_keys, alpha_5_9_master_keys,
                                divergent_keys, future_candidate_keys)
    build_needs_master_review_sheet(wb, needs_master_rows, master_data)
    build_needs_legal_review_sheet(wb, needs_legal_rows)
    build_divergent_sheet(wb, master_data, divergent_keys)
    build_future_candidate_sheet(wb, master_data, future_candidate_keys)
    build_valid_values_sheet(wb)
    build_gate_instructions_sheet(wb)

    wb.save(WORKBOOK_PATH)

    # 補助ファイル
    write_export_template_csv()

    manifest = {
        "version": PACKET_VERSION,
        "base_commit": ALPHA_5_10_BASE_COMMIT,
        "input_packet_version": INPUT_PACKET_VERSION,
        "input_gate_version": INPUT_GATE_VERSION,
        "generated_at": GENERATED_AT,
        "generator_script": "scripts/generate_alpha5_11_reviewer_workbook.py",
        "export_script": "scripts/export_alpha5_11_workbook_decisions.py",
        "scope": "internal_only",
        "purpose": "alpha.5.9 / alpha.5.10 を前提とした reviewer Excel ワークブック。master JSON は改変しない。",
        "workbook_created": True,
        "export_script_created": True,
        "public_release": False,
        "checked_promotion": False,
        "master_auto_update": False,
        "r8_provisional_used_for_checked": False,
        "release_pack_modified": False,
        "total_review_rows": len(decision_template_rows),
        "needs_master_review_count": len(needs_master_rows),
        "needs_legal_review_count": len(needs_legal_rows),
        "divergent_count": len(divergent_keys),
        "future_candidate_count": len(future_candidate_keys),
        "workbook_sheets": [
            "README", "Decision_Input_All",
            "Needs_Master_Review", "Needs_Legal_Review",
            "Divergent", "Future_Candidate",
            "Valid_Values", "Gate_Instructions",
        ],
        "valid_reviewer_decisions": list(VALID_DECISIONS),
        "modifying_decisions": sorted(MODIFYING_DECISIONS),
        "non_modifying_decisions": sorted(NON_MODIFYING_DECISIONS),
        "high_risk_decisions": sorted(HIGH_RISK_DECISIONS),
        "valid_implementation_allowed": list(VALID_IMPL_ALLOWED),
        "required_fields_when_yes": list(REQUIRED_FIELDS_WHEN_YES),
        "files": [
            "README.md",
            "reviewer_handoff_guide.md",
            "alpha5_11_reviewer_decision_workbook.xlsx",
            "reviewer_decision_export_template.csv",
            "workbook_export_instructions.md",
            "alpha5_11_reviewer_handoff_manifest.json",
        ],
        "invariants": [
            "master JSON 自動修正なし",
            "新規 checked 昇格なし",
            "R8.6 案資料を checked 昇格に使わない",
            "public release pack 未変更",
            "alpha.5.9 packet ファイル未破壊",
            "alpha.5.10 gate ファイル未破壊",
            "checked 20件 維持",
            "reviewer 入力ファイルを public に出さない",
        ],
    }
    (OUT_DIR / "alpha5_11_reviewer_handoff_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_readme(OUT_DIR, manifest)
    write_handoff_guide(OUT_DIR, manifest)
    write_export_instructions(OUT_DIR)

    print(f"alpha.5.11 reviewer handoff workbook generated at {OUT_DIR}")
    print(f"  workbook: {WORKBOOK_PATH.name}")
    print(f"  total_review_rows: {manifest['total_review_rows']}")
    print(f"  needs_master_review: {manifest['needs_master_review_count']}")
    print(f"  needs_legal_review : {manifest['needs_legal_review_count']}")
    print(f"  divergent          : {manifest['divergent_count']}")
    print(f"  future_candidate   : {manifest['future_candidate_count']}")


if __name__ == "__main__":
    main()
