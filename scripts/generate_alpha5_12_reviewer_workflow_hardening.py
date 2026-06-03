"""alpha.5.12 reviewer workflow hardening generator.

alpha.5.11 で作成した reviewer handoff workbook を、人間レビュー投入前に運用面で
強化する。

Phase 1: legal_review_clearance 列を追加（needs_legal_review 5件の手前ゲート）
Phase 2: Excel 備考欄 export 拡張（reviewer_role / review_note 等）
Phase 3: sample reviewed fixture を作って alpha.5.10 gate の分岐を全面検証
Phase 4: alpha.5.10 gate 拡張（本scriptとは別に gate 側で対応済）
Phase 5: workbook 更新（alpha.5.11 を破壊せず alpha.5.12 として別出力）
Phase 6: tests（test_reviewer_workflow_hardening.py で別途）

方針:
- master JSON は **絶対に修正しない**
- alpha.5.11 workbook / alpha.5.10 gate / alpha.5.9 packet は破壊しない
- public release pack は作らない
- 算定可否保証・公式コード完全照合済み・R8 改定対応完了 表現は禁止
"""
from __future__ import annotations

import csv
import io
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]

PACKET_VERSION = "alpha.5.12"
INPUT_PACKET_VERSION = "alpha.5.11"
INPUT_GATE_VERSION = "alpha.5.10"
ALPHA_5_11_BASE_COMMIT = "a3af77843f01653f1e7e10c99ece00b98faa2aba"
GENERATED_AT = "2026-05-10"

# Paths
ALPHA_5_9_PACKET_DIR = ROOT / "out" / "internal" / "alpha5_9_master_review_packet"
ALPHA_5_10_GATE_DIR = ROOT / "out" / "internal" / "alpha5_10_reviewer_decision_gate"
ALPHA_5_11_HANDOFF_DIR = ROOT / "out" / "internal" / "alpha5_11_reviewer_handoff_workbook"
ALPHA_5_11_WORKBOOK = ALPHA_5_11_HANDOFF_DIR / "alpha5_11_reviewer_decision_workbook.xlsx"
OUT_DIR = ROOT / "out" / "internal" / "alpha5_12_reviewer_workflow_hardening"
WORKBOOK_PATH = OUT_DIR / "alpha5_12_reviewer_decision_workbook.xlsx"
SAMPLE_INPUT_CSV = OUT_DIR / "sample_reviewed_decisions.csv"

GATE_SCRIPT = ROOT / "scripts" / "generate_alpha5_10_reviewer_decision_gate.py"

# Decision sets
VALID_DECISIONS = [
    "approve_official_code_addition", "keep_legacy_detection_only",
    "add_receipt_alias", "correct_internal_legacy_code",
    "mark_structural_mismatch", "keep_pattern_based_unverified",
    "escalate_legal_review", "defer_until_r8_definitive",
]
MODIFYING_DECISIONS = {
    "approve_official_code_addition", "add_receipt_alias",
    "correct_internal_legacy_code",
}
HIGH_RISK_DECISIONS = {"correct_internal_legacy_code"}
VALID_IMPL_ALLOWED = ["yes", "no", "pending"]
VALID_LEGAL_CLEARANCE = ["cleared", "not_cleared", "pending", "not_required"]
VALID_REVIEWER_ROLE = ["business_reviewer", "legal_reviewer", "final_approver"]
VALID_IMPL_PRIORITY = ["high", "medium", "low", "defer"]
VALID_RISK_ACK = ["yes", "no", "pending"]

# alpha.5.12 拡張 16列スキーマ
EXTENDED_COLUMNS = [
    "service", "kasan_key", "reviewer_decision",
    "reason", "required_evidence",
    "reviewer_name", "reviewed_at", "final_approved_by",
    "implementation_allowed",
    # alpha.5.12 拡張列
    "reviewer_role",
    "review_note",
    "legal_review_clearance",
    "legal_review_reference",
    "legal_review_note",
    "implementation_priority",
    "implementation_risk_acknowledged",
]

REQUIRED_FIELDS_WHEN_YES = (
    "reviewer_decision", "reason", "required_evidence",
    "reviewer_name", "reviewed_at", "final_approved_by",
)

# Decision_Input_All シートで使う列順 (workbook display)
DECISION_INPUT_DISPLAY_COLUMNS = [
    "service", "kasan_key", "kasan_display_name",
    "bucket", "current_overall_mapping_status", "proposed_action",
    "reviewer_decision", "reason", "required_evidence",
    "reviewer_name", "reviewed_at", "final_approved_by",
    "implementation_allowed",
    # alpha.5.12 拡張列（workbook 表示）
    "reviewer_role",
    "review_note",
    "legal_review_clearance",
    "legal_review_reference",
    "legal_review_note",
    "implementation_priority",
    "implementation_risk_acknowledged",
    "input_status_hint",
]


# ============================================================
# Style helpers
# ============================================================

def _border():
    side = Side(style="thin", color="888888")
    return Border(left=side, right=side, top=side, bottom=side)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="F8CBAD")
HIGH_RISK_FILL = PatternFill("solid", fgColor="F4B084")
NEEDS_LEGAL_FILL = PatternFill("solid", fgColor="DDEBF7")
FUTURE_CANDIDATE_FILL = PatternFill("solid", fgColor="E2EFDA")
DIVERGENT_FILL = PatternFill("solid", fgColor="FCE4D6")
REQUIRED_FIELD_FILL = PatternFill("solid", fgColor="FFE699")
LEGAL_CLEARANCE_FILL = PatternFill("solid", fgColor="B4C7E7")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def _set_header(ws, row_idx, columns):
    for ci, col_name in enumerate(columns, start=1):
        c = ws.cell(row=row_idx, column=ci, value=col_name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()


def _autosize(ws, max_width=80, min_width=8):
    for col_cells in ws.columns:
        first = col_cells[0]
        col_letter = get_column_letter(first.column)
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            for line in str(v).split("\n"):
                w = sum(2 if ord(c) > 127 else 1 for c in line)
                if w > max_len:
                    max_len = w
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


# ============================================================
# Loaders
# ============================================================

def _load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [r for r in reader if r.get("service") and not r.get("service", "").startswith("#")]


def load_master_kasans() -> dict:
    services = ("houmon_kango_kaigo", "tsusho_kaigo", "houmon_kaigo", "kyotaku_shien")
    out = {}
    for svc in services:
        path = ROOT / "regulatory_master" / "kaigo" / f"{svc}.json"
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        for k, v in (d.get("kasans") or {}).items():
            out[(svc, k)] = v
    return out


# ============================================================
# Sample fixture
# ============================================================

def build_sample_reviewed_decisions() -> list[dict]:
    """alpha.5.10 gate の全分岐を網羅する 12行のサンプル fixture を組み立てる。

    シナリオ:
    1. approved (master_review modifying): tsusho/chujudosha + add_receipt_alias
    2. approved (master_review modifying): tsusho/nyuyoku_II + approve_official_code_addition
    3. blocked (invalid decision): tsusho/koukuu_kinou_I + INVALID_DECISION
    4. blocked (missing required field): tsusho/eiyou_kaizen + impl=yes 要件不揃い
    5. blocked (high-risk no ack): houmon_kaigo/shokai_kasan + correct_internal_legacy_code 高リスク
    6. pending (blank): houmon_kaigo/seikatsu_kinou_renkei_I 完全空欄
    7. pending (impl=pending): houmon_kaigo/seikatsu_kinou_renkei_II decision あり / impl=pending
    8. legal_review_required (clearance=pending): houmon_kango/fukusu_mei_I_under30 + escalate
    9. legal_review_required (no clearance): houmon_kango/fukusu_mei_I_over30 + mark_structural
    10. approved (legal cleared): houmon_kango/fukusu_mei_II_under30 + approve / clearance=cleared
    11. blocked (future_candidate non-defer): houmon_kaigo/shougu_kaizen_kasan + approve
    12. pending (future_candidate defer): kyotaku_shien/shougu_kaizen_kasan_2026_06 + defer
    """
    blank_row = {col: "" for col in EXTENDED_COLUMNS}

    rows = []

    # 1. approved (modifying, low-risk)
    r = dict(blank_row, service="tsusho_kaigo", kasan_key="chujudosha_care_taisei",
             reviewer_decision="add_receipt_alias",
             reason="公式コード 155306 と社内コード 156271 の単位は一致するため alias 登録で対応",
             required_evidence="WAM_R7_8_DEFINITIVE_2025_03_28 PDF p131 で確認",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             review_note="alias 登録で PDF 検出を保ちつつ公式コード照合を充足できる",
             legal_review_clearance="not_required",
             implementation_priority="medium",
             implementation_risk_acknowledged="yes")
    rows.append(r)

    # 2. approved (modifying, low-risk)
    r = dict(blank_row, service="tsusho_kaigo", kasan_key="nyuyoku_II",
             reviewer_decision="approve_official_code_addition",
             reason="公式コード 155303 を master に追加",
             required_evidence="WAM_R7_8_DEFINITIVE PDF p131",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             review_note="単位差はなくコードのみ訂正",
             legal_review_clearance="not_required",
             implementation_priority="low",
             implementation_risk_acknowledged="yes")
    rows.append(r)

    # 3. blocked (invalid decision)
    r = dict(blank_row, service="tsusho_kaigo", kasan_key="koukuu_kinou_I",
             reviewer_decision="DEFINITELY_INVALID_DECISION",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             implementation_allowed="yes",
             reviewer_role="business_reviewer")
    rows.append(r)

    # 4. blocked (missing required_evidence when impl=yes)
    r = dict(blank_row, service="tsusho_kaigo", kasan_key="eiyou_kaizen",
             reviewer_decision="approve_official_code_addition",
             reason="公式コードに合わせる",
             required_evidence="",  # missing
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             implementation_risk_acknowledged="yes")
    rows.append(r)

    # 5. blocked (high-risk decision without risk_ack=yes)
    r = dict(blank_row, service="houmon_kaigo", kasan_key="shokai_kasan",
             reviewer_decision="correct_internal_legacy_code",
             reason="社内コード 116200 を公式 114001 に置換",
             required_evidence="WAM_R7_8_DEFINITIVE PDF",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             review_note="高リスクだが risk_ack 未確認のため blocked になる想定",
             legal_review_clearance="not_required",
             implementation_priority="medium",
             implementation_risk_acknowledged="no")  # blocked
    rows.append(r)

    # 6. pending (blank - 完全空欄)
    r = dict(blank_row, service="houmon_kaigo", kasan_key="seikatsu_kinou_renkei_I")
    rows.append(r)

    # 7. pending (impl=pending)
    r = dict(blank_row, service="houmon_kaigo", kasan_key="seikatsu_kinou_renkei_II",
             reviewer_decision="approve_official_code_addition",
             reason="検討中",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             implementation_allowed="pending",
             reviewer_role="business_reviewer")
    rows.append(r)

    # 8. legal_review_required (clearance=pending)
    r = dict(blank_row, service="houmon_kango_kaigo",
             kasan_key="fukusu_mei_houmon_kango_kasan_I_under30",
             reviewer_decision="escalate_legal_review",
             reason="基本コードへの付加加算構造の解釈確認が必要",
             required_evidence="老企第36号 解釈通知の確認待ち",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             legal_review_clearance="pending",
             implementation_risk_acknowledged="yes")
    rows.append(r)

    # 9. legal_review_required (no clearance)
    r = dict(blank_row, service="houmon_kango_kaigo",
             kasan_key="fukusu_mei_houmon_kango_kasan_I_over30",
             reviewer_decision="mark_structural_mismatch",
             reason="独立コードなし・structural として明示",
             required_evidence="現行通知では明確な独立コード未確認",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             implementation_risk_acknowledged="yes")
    # legal_review_clearance を空欄のままにする → legal_review_required 行き
    rows.append(r)

    # 10. approved (legal cleared)
    r = dict(blank_row, service="houmon_kango_kaigo",
             kasan_key="fukusu_mei_houmon_kango_kasan_II_under30",
             reviewer_decision="approve_official_code_addition",
             reason="法令確認者の clearance を経て公式コード 134200 として登録",
             required_evidence="令和8年法令解釈通知 (sample reference) を法令確認者が確認",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             review_note="法令確認者から clearance=cleared を取得済（sample fixture）",
             legal_review_clearance="cleared",
             legal_review_reference="令和8年5月XX日 老企第XXX号 (sample reference)",
             legal_review_note="法令確認者: sample_法令確認者B が解釈通知を確認",
             implementation_priority="high",
             implementation_risk_acknowledged="yes")
    rows.append(r)

    # 11. blocked (future_candidate_only に approve を入れた)
    r = dict(blank_row, service="houmon_kaigo", kasan_key="shougu_kaizen_kasan",
             reviewer_decision="approve_official_code_addition",
             reason="本来は defer すべき",
             required_evidence="X", reviewer_name="X", reviewed_at="2026-05-15",
             final_approved_by="X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             implementation_risk_acknowledged="yes")
    rows.append(r)

    # 12. pending (future_candidate_only with defer)
    r = dict(blank_row, service="kyotaku_shien", kasan_key="shougu_kaizen_kasan_2026_06",
             reviewer_decision="defer_until_r8_definitive",
             reason="R8.6.1 確定版が出るまで保留",
             required_evidence="(R8.6.1 確定版PDFが出るまで)",
             reviewer_name="sample_業務担当A", reviewed_at="2026-05-15",
             final_approved_by="sample_最終承認者X",
             implementation_allowed="yes",
             reviewer_role="business_reviewer",
             review_note="R8.6.1 確定版 PDF 公開前は defer",
             legal_review_clearance="not_required",
             implementation_priority="defer",
             implementation_risk_acknowledged="yes")
    rows.append(r)

    return rows


def write_sample_input_csv(rows: list[dict]):
    with open(SAMPLE_INPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=EXTENDED_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
        f.write("\n")
        f.write(f"# alpha.5.12 sample reviewed decisions fixture\n")
        f.write(f"# generated_at: {GENERATED_AT}\n")
        f.write("# 12 行で alpha.5.10 gate の全分岐を網羅\n")
        f.write("# 各値は架空・公開デモ用。実 reviewer 入力ではない。\n")


# ============================================================
# Workbook builders
# ============================================================

def build_readme_sheet(wb):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("alpha.5.12 Reviewer Decision Workbook (workflow hardening edition)", Font(bold=True)),
        ("", None),
        (f"version: {PACKET_VERSION}", None),
        (f"base_commit: {ALPHA_5_11_BASE_COMMIT} (alpha.5.11)", None),
        (f"input_packet_version: {INPUT_PACKET_VERSION}", None),
        (f"input_gate_version: {INPUT_GATE_VERSION}", None),
        (f"generated_at: {GENERATED_AT}", None),
        ("", None),
        ("【alpha.5.11 → alpha.5.12 拡張点】", Font(bold=True)),
        ("- legal_review_clearance / legal_review_reference / legal_review_note 列を追加", None),
        ("- reviewer_role / review_note / implementation_priority / implementation_risk_acknowledged 列を追加", None),
        ("- alpha.5.10 gate を拡張: legal cleared 判定・high-risk decision の risk_ack=yes 必須化", None),
        ("- 後方互換: legacy 9列CSV も従来どおり読める", None),
        ("- alpha.5.11 workbook は破壊しない（別ファイル alpha5_12_reviewer_decision_workbook.xlsx として出力）", None),
        ("", None),
        ("【public 資料ではありません】", Font(bold=True)),
        ("- このブックは out/internal 配下の内部レビュー資料", None),
        ("- reviewer が入力した実判断ファイルは public に出さない", None),
        ("- alpha.5.3 / alpha.5.4 release pack には影響しない", None),
        ("", None),
        ("【入力しても自動で master 修正されません】", Font(bold=True)),
        ("- このブックへの入力は master JSON を改変しません", None),
        ("- export script で CSV 化したあと、alpha.5.10 gate で検証してから次フェーズに進みます", None),
        ("- approved 候補に入っても master JSON は alpha.5.13+ で別 PR にて段階反映", None),
        ("", None),
        ("【シート一覧】", Font(bold=True)),
        ("- README                : 本ファイル", None),
        ("- Decision_Input_All    : 入力シート（38件・拡張列を含む 21列・プルダウン・色分け）", None),
        ("- Needs_Master_Review   : 業務担当向け参照（28件）", None),
        ("- Needs_Legal_Review    : 法令確認者向け参照（5件・legal_review_clearance 列を追加）", None),
        ("- Divergent             : divergent 3件（参照のみ）", None),
        ("- Future_Candidate      : R8.6.1 確定版待ち2件（必ず defer）", None),
        ("- Valid_Values          : 全選択肢（alpha.5.12 拡張カテゴリ込み）", None),
        ("- Gate_Instructions     : Excel入力後の export → gate 再実行手順（alpha.5.12 版）", None),
    ]
    for r_idx, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=r_idx, column=1, value=text)
        if font:
            c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    return ws


def build_valid_values_sheet(wb):
    ws = wb.create_sheet("Valid_Values")
    columns = ["category", "value", "is_modifying", "is_high_risk", "note"]
    _set_header(ws, 1, columns)
    rows = []
    for d in VALID_DECISIONS:
        is_mod = d in MODIFYING_DECISIONS
        is_hr = d in HIGH_RISK_DECISIONS
        note = {
            "approve_official_code_addition": "公式コードを追加。低リスク。",
            "keep_legacy_detection_only": "社内 legacy code 維持。master 修正なし。",
            "add_receipt_alias": "公式コードを社内コードの alias 登録。中リスク。",
            "correct_internal_legacy_code": "社内 service_codes を公式コードに置換。**高リスク**（risk_ack=yes 必須・PDF検出回帰必須）。",
            "mark_structural_mismatch": "structural_mismatch として明示。法令解釈待ち。",
            "keep_pattern_based_unverified": "パターン検出のまま継続。master 修正なし。",
            "escalate_legal_review": "法令確認者にエスカレーション。approved にならない。",
            "defer_until_r8_definitive": "R8.6.1 確定版が出るまで保留。",
        }.get(d, "")
        rows.append(("reviewer_decision", d, "yes" if is_mod else "no",
                     "yes" if is_hr else "no", note))
    for v in VALID_IMPL_ALLOWED:
        note = {
            "yes": "実装してよい。**6 必須フィールド全揃い** + 高リスク decision の場合は implementation_risk_acknowledged=yes も必須",
            "no": "実装しない。決定のみ記録。",
            "pending": "保留。決定確定前。",
        }[v]
        rows.append(("implementation_allowed", v, "", "", note))
    # alpha.5.12 拡張カテゴリ
    for v in VALID_LEGAL_CLEARANCE:
        note = {
            "cleared": "法令確認者が clearance を付与済。**他の必須フィールド全揃い + impl=yes** で approved 候補に進める。",
            "not_cleared": "法令確認の結果、clearance が下りなかった。approved にならない。",
            "pending": "法令確認者の確認中。approved にならない。",
            "not_required": "needs_legal_review に該当しない加算で reviewer が便宜的に記載。判定には使わない。",
        }[v]
        rows.append(("legal_review_clearance", v, "", "", note))
    for v in VALID_REVIEWER_ROLE:
        note = {
            "business_reviewer": "業務担当（needs_master_review / divergent の判断）",
            "legal_reviewer": "法令確認者（needs_legal_review の clearance 判定）",
            "final_approver": "最終承認者（implementation_allowed=yes の最終ハンコ）",
        }[v]
        rows.append(("reviewer_role", v, "", "", note))
    for v in VALID_IMPL_PRIORITY:
        note = {
            "high": "高優先度（早期実装）",
            "medium": "中優先度",
            "low": "低優先度",
            "defer": "保留・将来再評価",
        }[v]
        rows.append(("implementation_priority", v, "", "", note))
    for v in VALID_RISK_ACK:
        note = {
            "yes": "高リスク decision を実施するリスクを reviewer が確認済（PDF検出回帰テスト・段階的 PR 等）",
            "no": "リスク認識なし → 高リスク decision の場合は **blocked**",
            "pending": "リスク認識中（最終判断前）",
        }[v]
        rows.append(("implementation_risk_acknowledged", v, "", "", note))
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            if val in HIGH_RISK_DECISIONS:
                c.fill = HIGH_RISK_FILL
    _autosize(ws)
    return ws


def build_decision_input_sheet(wb, decision_template_rows, master_data,
                                 alpha_5_9_legal_keys, divergent_keys, future_candidate_keys):
    ws = wb.create_sheet("Decision_Input_All")
    columns = DECISION_INPUT_DISPLAY_COLUMNS
    _set_header(ws, 1, columns)

    note_text = (
        "ガイダンス（alpha.5.12 hardening edition）: "
        "implementation_allowed=yes の場合は reviewer_decision / reason / required_evidence / "
        "reviewer_name / reviewed_at / final_approved_by が必須（黄色）。"
        "needs_legal_review 5件は legal_review_clearance=cleared + legal_review_reference あり + "
        "他必須フィールド全揃い + impl=yes でないと approved にならない（薄水色）。"
        "future_candidate_only 2件は defer_until_r8_definitive のみ可（薄緑）。"
        "correct_internal_legacy_code は高リスクで implementation_risk_acknowledged=yes が必須（濃オレンジ）。"
    )
    note_cell = ws.cell(row=2, column=1, value=note_text)
    note_cell.fill = NOTE_FILL
    note_cell.alignment = WRAP_ALIGN
    note_cell.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    start_data_row = 3
    for ri, row in enumerate(decision_template_rows, start=start_data_row):
        svc = row.get("service", "")
        k = row.get("kasan_key", "")
        kasan_def = master_data.get((svc, k)) or {}
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        pa = three.get("proposed_action", "")
        ov = kasan_def.get("overall_mapping_status", "")
        if (svc, k) in alpha_5_9_legal_keys:
            bucket = "needs_legal_review"
        elif pa == "future_candidate_only":
            bucket = "future_candidate_only"
        elif (svc, k) in divergent_keys:
            bucket = "divergent (keep_pattern_based_unverified ∧ overall=needs_review)"
        elif pa == "needs_master_review":
            bucket = "needs_master_review"
        else:
            bucket = pa or ""
        hint_parts = []
        if bucket == "needs_legal_review":
            hint_parts.append("⚠ legal review 必須・clearance=cleared が無いと approved にならない")
        if bucket == "future_candidate_only":
            hint_parts.append("⚠ R8.6.1 確定版待ち・defer_until_r8_definitive のみ可")
        if bucket.startswith("divergent"):
            hint_parts.append("⚠ divergent 加算・alpha.5.8.1 で audit_note 化済")
        hint = " / ".join(hint_parts) or "業務担当が判断・記入"

        cells_data = [
            svc, k, kasan_def.get("name", ""),
            bucket, ov, pa,
            row.get("reviewer_decision", ""),
            row.get("reason", ""),
            row.get("required_evidence", ""),
            row.get("reviewer_name", ""),
            row.get("reviewed_at", ""),
            row.get("final_approved_by", ""),
            row.get("implementation_allowed", ""),
            row.get("reviewer_role", ""),
            row.get("review_note", ""),
            row.get("legal_review_clearance", ""),
            row.get("legal_review_reference", ""),
            row.get("legal_review_note", ""),
            row.get("implementation_priority", ""),
            row.get("implementation_risk_acknowledged", ""),
            hint,
        ]
        for ci, val in enumerate(cells_data, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()

        if bucket == "needs_legal_review":
            for ci in range(1, len(columns) + 1):
                ws.cell(row=ri, column=ci).fill = NEEDS_LEGAL_FILL
        elif bucket == "future_candidate_only":
            for ci in range(1, len(columns) + 1):
                ws.cell(row=ri, column=ci).fill = FUTURE_CANDIDATE_FILL
        elif bucket.startswith("divergent"):
            for ci in range(1, len(columns) + 1):
                ws.cell(row=ri, column=ci).fill = DIVERGENT_FILL

        # required-field 黄色マーキング (col 7-12, 13)
        for col_idx in (7, 8, 9, 10, 11, 12, 13):
            c = ws.cell(row=ri, column=col_idx)
            if c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = REQUIRED_FIELD_FILL
        # legal_review_clearance (col 16) は薄水系で強調（needs_legal_review でない場合）
        legal_cell = ws.cell(row=ri, column=16)
        if bucket != "needs_legal_review" and legal_cell.fill.fgColor.rgb in (None, "00000000"):
            legal_cell.fill = LEGAL_CLEARANCE_FILL

    n_data_rows = len(decision_template_rows)
    last_row = start_data_row + n_data_rows - 1

    # データ検証 (プルダウン)
    validations = [
        ("G", VALID_DECISIONS, "無効な reviewer_decision",
         "Valid_Values シートの reviewer_decision から選択してください"),
        ("M", VALID_IMPL_ALLOWED, "無効な implementation_allowed",
         "yes / no / pending のいずれかを選択してください"),
        ("N", VALID_REVIEWER_ROLE, "無効な reviewer_role",
         "business_reviewer / legal_reviewer / final_approver から選択"),
        ("P", VALID_LEGAL_CLEARANCE, "無効な legal_review_clearance",
         "cleared / not_cleared / pending / not_required から選択"),
        ("S", VALID_IMPL_PRIORITY, "無効な implementation_priority",
         "high / medium / low / defer から選択"),
        ("T", VALID_RISK_ACK, "無効な implementation_risk_acknowledged",
         "yes / no / pending から選択"),
    ]
    for col_letter, choices, err_title, err_msg in validations:
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(choices) + '"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=err_title,
            error=err_msg,
        )
        dv.add(f"{col_letter}{start_data_row}:{col_letter}{last_row}")
        ws.add_data_validation(dv)

    ws.freeze_panes = f"D{start_data_row}"
    _autosize(ws, max_width=40)
    return ws


def build_needs_master_review_sheet(wb, rows):
    ws = wb.create_sheet("Needs_Master_Review")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "official_service_code", "official_name", "official_unit", "official_calc_unit",
        "official_source_id",
        "receipt_detection_code", "receipt_detection_pattern", "receipt_detection_status",
        "internal_legacy_code", "internal_legacy_unit",
        "mismatch_type", "mismatch_summary",
        "proposed_review_question", "recommended_next_step",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(row=2, column=1, value="（業務担当向け）公式コードと社内 legacy コードの照合を行ってください。decision を入力するのは Decision_Input_All シートです。")
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    for ri, row in enumerate(rows, start=3):
        cells = [row.get(c, "") for c in columns]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()

    ws.freeze_panes = "D3"
    _autosize(ws, max_width=50)
    return ws


def build_needs_legal_review_sheet(wb, rows):
    ws = wb.create_sheet("Needs_Legal_Review")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "legal_review_reason", "structural_issue_type",
        "official_source_id", "internal_legacy_code", "receipt_detection_pattern",
        "current_overall_mapping_status",
        "why_not_checked", "legal_question", "reference_needed",
        "recommended_next_step",
        # alpha.5.12 拡張列
        "legal_review_clearance",
        "legal_review_reference",
        "legal_review_note",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(
        row=2, column=1,
        value=("（法令確認者向け）alpha.5.12 拡張: legal_review_clearance / legal_review_reference / "
               "legal_review_note を本シートで記録できます。Decision_Input_All の同名列にも転記してください。"
               " clearance=cleared + reference あり + impl=yes + 必須フィールド全揃いの場合のみ "
               "approved 候補に進めます（gate 側で判定）。"))
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    for ri, row in enumerate(rows, start=3):
        cells = [
            row.get("service", ""), row.get("kasan_key", ""), row.get("kasan_display_name", ""),
            row.get("legal_review_reason", ""), row.get("structural_issue_type", ""),
            row.get("official_source_id", ""), row.get("internal_legacy_code", ""),
            row.get("receipt_detection_pattern", ""),
            row.get("current_overall_mapping_status", ""),
            row.get("why_not_checked", ""), row.get("legal_question", ""),
            row.get("reference_needed", ""), row.get("recommended_next_step", ""),
            "", "", "",  # legal_review_clearance / reference / note
        ]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            c.fill = NEEDS_LEGAL_FILL
        # alpha.5.12 拡張列を強調（薄水色 → さらに濃く）
        for col_idx in (14, 15, 16):
            ws.cell(row=ri, column=col_idx).fill = LEGAL_CLEARANCE_FILL

    # legal_review_clearance プルダウン
    n_rows = len(rows)
    last_row = 2 + n_rows
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(VALID_LEGAL_CLEARANCE) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="無効な legal_review_clearance",
        error="cleared / not_cleared / pending / not_required から選択",
    )
    dv.add(f"N3:N{last_row}")
    ws.add_data_validation(dv)

    ws.freeze_panes = "D3"
    _autosize(ws, max_width=50)
    return ws


def build_divergent_sheet(wb, master_data, divergent_keys):
    ws = wb.create_sheet("Divergent")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "proposed_action", "overall_mapping_status",
        "official_match_type", "official_code", "internal_legacy_code",
        "divergence_reason",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(row=2, column=1, value="（参照のみ）alpha.5.8.1 で audit_note 化済の divergent 3件。")
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    ri = 3
    for key in sorted(divergent_keys):
        kasan_def = master_data.get(key)
        if not kasan_def:
            continue
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        official = three.get("official_code_model") or {}
        legacy = three.get("internal_legacy_model") or {}
        note_obj = three.get("alpha_5_8_1_proposed_overall_divergence_note") or {}
        cells = [
            key[0], key[1], kasan_def.get("name", ""),
            three.get("proposed_action", ""),
            kasan_def.get("overall_mapping_status", ""),
            official.get("official_match_type", ""),
            official.get("official_service_code") or "",
            legacy.get("internal_legacy_code") or "",
            note_obj.get("reason", ""),
        ]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            c.fill = DIVERGENT_FILL
        ri += 1
    ws.freeze_panes = "D3"
    _autosize(ws, max_width=80)
    return ws


def build_future_candidate_sheet(wb, master_data, future_candidate_keys):
    ws = wb.create_sheet("Future_Candidate")
    columns = [
        "service", "kasan_key", "kasan_display_name",
        "proposed_action", "overall_mapping_status",
        "official_source_id", "audit_note", "required_decision",
    ]
    _set_header(ws, 1, columns)
    note = ws.cell(
        row=2, column=1,
        value=("（参照のみ）future_candidate_only 2件。R8.6.1 確定版が出るまで "
               "**reviewer_decision は defer_until_r8_definitive のみ**受理。"
               " R8.6 案資料 (WAM_R8_6_8_PROVISIONAL_2026_04_30) は checked 昇格に使わない。"))
    note.fill = NOTE_FILL
    note.alignment = WRAP_ALIGN
    note.font = Font(bold=True, color="9C5700")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    ri = 3
    for key in sorted(future_candidate_keys):
        kasan_def = master_data.get(key)
        if not kasan_def:
            continue
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        official = three.get("official_code_model") or {}
        cells = [
            key[0], key[1], kasan_def.get("name", ""),
            three.get("proposed_action", ""),
            kasan_def.get("overall_mapping_status", ""),
            official.get("source_id") or "",
            three.get("audit_note", ""),
            "defer_until_r8_definitive",
        ]
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP_ALIGN
            c.border = _border()
            c.fill = FUTURE_CANDIDATE_FILL
        ri += 1
    ws.freeze_panes = "D3"
    _autosize(ws, max_width=80)
    return ws


def build_gate_instructions_sheet(wb):
    ws = wb.create_sheet("Gate_Instructions")
    blocks = [
        ("Excel 入力後の流れ (alpha.5.12 hardening edition)", Font(bold=True)),
        ("", None),
        ("1. このブックを保存（.xlsx 形式・上書き）", None),
        ("2. ターミナルでカレントを products/kasan-manager に移動", None),
        ("3. 以下を実行（alpha.5.12 拡張列を含む 16列CSVが出る）:", None),
        ("   python scripts/export_alpha5_11_workbook_decisions.py \\", None),
        ("     --workbook out/internal/alpha5_12_reviewer_workflow_hardening/alpha5_12_reviewer_decision_workbook.xlsx \\", None),
        ("     --output out/internal/alpha5_12_reviewer_workflow_hardening/reviewer_decision_export.csv", None),
        ("", None),
        ("alpha.5.10 gate 拡張版（legal cleared / high-risk 判定込み）の使い方:", Font(bold=True)),
        ("python scripts/generate_alpha5_10_reviewer_decision_gate.py \\", None),
        ("  --input out/internal/alpha5_12_reviewer_workflow_hardening/reviewer_decision_export.csv \\", None),
        ("  --output out/internal/alpha5_10_reviewer_decision_gate_from_alpha5_12_workbook/", None),
        ("", None),
        ("alpha.5.10 gate は CSV header から拡張列の有無を自動検出します。legacy 9列CSV も従来どおり読めます。", None),
        ("", None),
        ("alpha.5.12 で追加された分類ルール:", Font(bold=True)),
        ("- needs_legal_review 行は legal_review_clearance=cleared + legal_review_reference あり + impl=yes + 必須揃い → approved 候補", None),
        ("- 上記が揃わない場合 → legal_review_required（従来挙動）", None),
        ("- correct_internal_legacy_code (高リスク) は implementation_risk_acknowledged=yes が必須 → 無ければ blocked", None),
        ("- future_candidate_only は legal_review_clearance があっても approved にしない（必ず defer）", None),
        ("", None),
        ("blocked_reason 凡例（alpha.5.12 で追加）:", Font(bold=True)),
        ("- high_risk_decision_requires_implementation_risk_acknowledged_yes", None),
        ("- invalid_legal_review_clearance / invalid_implementation_risk_acknowledged", None),
        ("- needs_legal_review_kasan_pending_legal_clearance（従来通り）", None),
        ("- future_candidate_only_must_be_defer_until_r8_definitive（従来通り）", None),
    ]
    for r_idx, (text, font) in enumerate(blocks, start=1):
        c = ws.cell(row=r_idx, column=1, value=text)
        if font:
            c.font = font
        c.alignment = WRAP_ALIGN
    ws.column_dimensions["A"].width = 130
    return ws


# ============================================================
# Run gate against sample fixture and copy outputs with sample_ prefix
# ============================================================

def run_gate_on_sample(sample_csv: Path, out_dir: Path):
    """alpha.5.10 gate を sample CSV に対して実行し、出力を sample_ prefix で out_dir にコピーする。"""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_out = Path(tmpdir) / "gate_out"
        tmp_out.mkdir(parents=True, exist_ok=True)
        result = subprocess.run(
            [sys.executable, str(GATE_SCRIPT),
             "--input", str(sample_csv), "--output", str(tmp_out)],
            cwd=ROOT, capture_output=True, text=True,
        )
        if result.returncode != 0:
            raise SystemExit(f"gate failed: {result.stderr}")

        # 必要な出力を sample_ prefix で out_dir にコピー
        copy_map = {
            "decision_validation_report.md": "sample_reviewed_decision_validation_report.md",
            "approved_changes_preview.csv": "sample_approved_changes_preview.csv",
            "blocked_or_incomplete_decisions.csv": "sample_blocked_or_incomplete_decisions.csv",
            "pending_decisions.csv": "sample_pending_decisions.csv",
            "legal_review_required.csv": "sample_legal_review_required.csv",
        }
        for src_name, dst_name in copy_map.items():
            src = tmp_out / src_name
            dst = out_dir / dst_name
            if src.exists():
                shutil.copyfile(src, dst)

        # gate manifest を読んで件数を返す
        manifest_path = tmp_out / "alpha5_10_reviewer_decision_gate_manifest.json"
        gate_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        return gate_manifest


# ============================================================
# Aux docs
# ============================================================

def write_legal_clearance_rules(out_dir: Path):
    text = f"""# alpha.5.12 Legal Clearance Rules

**version**: {PACKET_VERSION}
**base_commit**: `{ALPHA_5_11_BASE_COMMIT}` (alpha.5.11)
**generated_at**: {GENERATED_AT}

---

## 概要

alpha.5.11 までは needs_legal_review 5件（複数名訪問看護加算 4件 + 長時間訪問看護加算 1件）は、
法令確認後も `legal_review_required` バケットに滞留する仕様でした。
alpha.5.12 では `legal_review_clearance` フラグを導入し、法令確認者がレビュー結果を記録できるようにします。

## legal_review_clearance の許可値

| 値 | 意味 | gate での扱い |
|---|---|---|
| `cleared` | 法令確認者が clearance を付与（解釈通知などを参照済み） | 他の必須フィールドが全揃いなら approved 候補に進める |
| `not_cleared` | 法令確認の結果、clearance が下りなかった | legal_review_required（approved にならない） |
| `pending` | 法令確認者がまだ確認中 | legal_review_required（approved にならない） |
| `not_required` | needs_legal_review に該当しない加算で reviewer が便宜的に記載 | 判定には使わない（needs_legal_review 以外の行で使用） |

## legal cleared → approved の必須条件

`needs_legal_review` バケットの行が approved 候補に進むには **以下の全条件** を満たす必要があります:

1. `legal_review_clearance == "cleared"`
2. `legal_review_reference` が空でない（解釈通知の番号や事務連絡日付など）
3. `implementation_allowed == "yes"`
4. `final_approved_by` が空でない（最終承認者）
5. `required_evidence` が空でない（PDF page など）
6. `reviewer_decision` が `MODIFYING_DECISIONS` に含まれる
   - approve_official_code_addition / add_receipt_alias / correct_internal_legacy_code
7. `reviewer_decision` が `correct_internal_legacy_code`（高リスク）の場合、
   `implementation_risk_acknowledged == "yes"` も必須

## サンプル: clearance あり approved 候補

```csv
service,kasan_key,reviewer_decision,reason,required_evidence,reviewer_name,reviewed_at,final_approved_by,implementation_allowed,reviewer_role,review_note,legal_review_clearance,legal_review_reference,legal_review_note,implementation_priority,implementation_risk_acknowledged
houmon_kango_kaigo,fukusu_mei_houmon_kango_kasan_II_under30,approve_official_code_addition,法令確認者の clearance を経て公式コード 134200 として登録,令和8年法令解釈通知 (sample reference) を法令確認者が確認,sample_業務担当A,2026-05-15,sample_最終承認者X,yes,business_reviewer,法令確認者から clearance=cleared を取得済,cleared,令和8年5月XX日 老企第XXX号 (sample reference),法令確認者: sample_法令確認者B が解釈通知を確認,high,yes
```

→ alpha.5.10 gate で **approved_changes_preview** に分類

## サンプル: clearance なし → legal_review_required

clearance が `pending` / `not_cleared` / 空欄のいずれかなら、impl=yes でも legal_review_required:

```csv
houmon_kango_kaigo,fukusu_mei_houmon_kango_kasan_I_under30,escalate_legal_review,...,yes,business_reviewer,...,pending,...
```

→ alpha.5.10 gate で **legal_review_required** に分類

## R8.6 案資料の扱い（変更なし）

- WAM_R8_6_8_PROVISIONAL_2026_04_30 は **案資料**で `checked_promotion_allowed=false`
- future_candidate_only 2 件 (訪介 shougu_kaizen_kasan / 居宅 shougu_kaizen_kasan_2026_06)
  は legal_review_clearance があっても **必ず defer_until_r8_definitive のみ受理**
- 他の decision を入れると alpha.5.10 gate で **blocked**

## 実装範囲（alpha.5.12 - alpha.5.13+）

- alpha.5.12: gate 側で legal_review_clearance を受け付け、approved 候補に流す
- alpha.5.12: master JSON は **改変しない**（候補提示のみ）
- **alpha.5.13+**: approved 候補を別 PR で master JSON に段階反映（一括反映禁止）

## 関連ファイル

- alpha5_12_reviewer_decision_workbook.xlsx — Decision_Input_All シートに legal_review_clearance 列
- sample_reviewed_decisions.csv — fixture 12 シナリオ（うち legal cleared 1件）
- alpha.5.10 gate — generate_alpha5_10_reviewer_decision_gate.py に判定ロジック追加済
"""
    (out_dir / "legal_clearance_rules.md").write_text(text, encoding="utf-8")


def write_readme(out_dir: Path, manifest: dict, gate_manifest: dict):
    text = f"""# alpha.5.12 Reviewer Workflow Hardening

**version**: {PACKET_VERSION}
**base_commit**: `{ALPHA_5_11_BASE_COMMIT}` (alpha.5.11)
**input_packet_version**: {INPUT_PACKET_VERSION}
**input_gate_version**: {INPUT_GATE_VERSION}
**generated_at**: {GENERATED_AT}

---

## 位置付け

alpha.5.11 で作成した reviewer handoff workbook を、人間レビュー投入前に運用面で強化したパケット。
特に、`legal_review_clearance` フラグと Excel 備考欄 export 拡張、サンプルレビュー fixture による
gate 受け入れテストを追加し、実レビュー後に alpha.5.10 gate が詰まらない状態にする。

- 出力先: `out/internal/alpha5_12_reviewer_workflow_hardening/`（**out/internal 配下のみ**・public release pack には含めない）
- alpha.5.11 workbook は **破壊しない**（別ファイル `alpha5_12_reviewer_decision_workbook.xlsx` として出力）
- alpha.5.10 gate / alpha.5.9 packet も破壊しない

## alpha.5.11 → alpha.5.12 拡張点

| 項目 | alpha.5.11 | alpha.5.12 |
|---|---|---|
| Decision_Input_All の列数 | 14列 | **21列** (拡張7列追加) |
| Needs_Legal_Review の列数 | 14列 | **16列** (legal_review_clearance / reference / note 追加) |
| Valid_Values のカテゴリ数 | 3 (decision/impl/role) | **6** (上記 + legal_clearance/priority/risk_ack) |
| export CSV のスキーマ | legacy 9列 | extended 16列 (auto detect) |
| alpha.5.10 gate | 9列専用 | **拡張16列も読める** (後方互換) |
| legal cleared 判定 | なし (滞留) | **あり** (clearance=cleared で approved 候補へ) |
| 高リスク decision | implementation_risk フラグなし | **risk_ack=yes 必須** (なければ blocked) |

## 不変条件 (テストで保護)

- ❌ master JSON 自動修正なし
- ❌ 新規 checked 昇格なし
- ❌ R8.6 案資料は checked 昇格に使わない
- ❌ public release pack は本 alpha.5.12 で更新しない
- ❌ alpha.5.9 packet / alpha.5.10 gate / alpha.5.11 workbook は破壊しない
- ❌ 過剰な完了感を与える表現は使わない（disclaimer 維持）
- ❌ reviewer 入力ファイルを public に出さない

## 含まれるファイル

| ファイル | 内容 |
|---|---|
| `README.md` | 本ファイル |
| `alpha5_12_reviewer_decision_workbook.xlsx` | 8シート構成のレビュー用ブック (alpha.5.12版) |
| `sample_reviewed_decisions.csv` | 12シナリオの sample fixture (extended 16列) |
| `sample_reviewed_decision_validation_report.md` | sample に対する gate の検証レポート |
| `sample_approved_changes_preview.csv` | sample に対する gate の approved 結果 |
| `sample_blocked_or_incomplete_decisions.csv` | sample に対する gate の blocked 結果 |
| `sample_pending_decisions.csv` | sample に対する gate の pending 結果 |
| `sample_legal_review_required.csv` | sample に対する gate の legal_review_required 結果 |
| `legal_clearance_rules.md` | legal_review_clearance のルール詳細 |
| `alpha5_12_reviewer_workflow_hardening_manifest.json` | パケットメタデータ |

## sample_reviewed_decisions.csv のシナリオ

| # | service | kasan_key | scenario | 想定 bucket |
|---|---|---|---|---|
| 1 | tsusho_kaigo | chujudosha_care_taisei | add_receipt_alias + impl=yes | approved |
| 2 | tsusho_kaigo | nyuyoku_II | approve_official_code_addition + impl=yes | approved |
| 3 | tsusho_kaigo | koukuu_kinou_I | INVALID_DECISION | blocked |
| 4 | tsusho_kaigo | eiyou_kaizen | impl=yes 必須欠落 | blocked |
| 5 | houmon_kaigo | shokai_kasan | correct_internal_legacy_code + risk_ack=no | blocked |
| 6 | houmon_kaigo | seikatsu_kinou_renkei_I | 完全空欄 | pending |
| 7 | houmon_kaigo | seikatsu_kinou_renkei_II | impl=pending | pending |
| 8 | houmon_kango_kaigo | fukusu_mei_houmon_kango_kasan_I_under30 | clearance=pending | legal_review_required |
| 9 | houmon_kango_kaigo | fukusu_mei_houmon_kango_kasan_I_over30 | clearance 空欄 | legal_review_required |
| 10 | houmon_kango_kaigo | fukusu_mei_houmon_kango_kasan_II_under30 | clearance=cleared + 全揃い | **approved (legal cleared)** |
| 11 | houmon_kaigo | shougu_kaizen_kasan | future_candidate に approve | blocked |
| 12 | kyotaku_shien | shougu_kaizen_kasan_2026_06 | future_candidate に defer | pending |

## sample に対する gate 結果

| バケット | 件数 |
|---|---:|
| approved | {gate_manifest['approved_count']} |
| blocked | {gate_manifest['blocked_count']} |
| pending | {gate_manifest['pending_count']} |
| legal_review_required | {gate_manifest['legal_review_required_count']} |
| **合計** | **{gate_manifest['approved_count'] + gate_manifest['blocked_count'] + gate_manifest['pending_count'] + gate_manifest['legal_review_required_count']}** |

## 再生成方法

```
cd products/kasan-manager
python scripts/generate_alpha5_12_reviewer_workflow_hardening.py
```

## reviewer 後の運用フロー（alpha.5.12 版）

1. reviewer が alpha5_12_reviewer_decision_workbook.xlsx を開く
2. Decision_Input_All シートで 38 行を埋める（拡張列含む）
3. needs_legal_review 5件は法令確認者が legal_review_clearance を埋める
4. 高リスク decision (correct_internal_legacy_code) は implementation_risk_acknowledged=yes を付ける
5. 保存後にターミナル: `python scripts/export_alpha5_11_workbook_decisions.py --workbook .../alpha5_12_reviewer_decision_workbook.xlsx --output .../reviewer_decision_export.csv`
6. ゲート再実行: `python scripts/generate_alpha5_10_reviewer_decision_gate.py --input .../reviewer_decision_export.csv --output .../alpha5_10_reviewer_decision_gate_from_alpha5_12_workbook/`
7. blocked / pending / legal_review_required を解消するまで 2〜6 を繰り返す
8. **alpha.5.13+ で別途 PR を立て、approved 行のみを段階的に master JSON へ反映**
"""
    (out_dir / "README.md").write_text(text, encoding="utf-8")


# ============================================================
# Main
# ============================================================

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # alpha.5.9 packet の reference data 読み込み
    decision_template_rows = _load_csv(ALPHA_5_9_PACKET_DIR / "reviewer_decision_template.csv")
    needs_master_rows = _load_csv(ALPHA_5_9_PACKET_DIR / "needs_master_review_matrix.csv")
    needs_legal_rows = _load_csv(ALPHA_5_9_PACKET_DIR / "needs_legal_review_matrix.csv")

    master_data = load_master_kasans()

    alpha_5_9_legal_keys = {(r["service"], r["kasan_key"]) for r in needs_legal_rows}
    divergent_keys = set()
    future_candidate_keys = set()
    for key, kasan_def in master_data.items():
        three = (kasan_def.get("service_code_audit") or {}).get("alpha_5_8_three_layer_model") or {}
        pa = three.get("proposed_action")
        ov = kasan_def.get("overall_mapping_status")
        if pa == "future_candidate_only":
            future_candidate_keys.add(key)
        if pa == "keep_pattern_based_unverified" and ov == "needs_review":
            divergent_keys.add(key)

    # ----- alpha.5.12 workbook 構築 -----
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    build_readme_sheet(wb)
    build_decision_input_sheet(wb, decision_template_rows, master_data,
                                 alpha_5_9_legal_keys, divergent_keys, future_candidate_keys)
    build_needs_master_review_sheet(wb, needs_master_rows)
    build_needs_legal_review_sheet(wb, needs_legal_rows)
    build_divergent_sheet(wb, master_data, divergent_keys)
    build_future_candidate_sheet(wb, master_data, future_candidate_keys)
    build_valid_values_sheet(wb)
    build_gate_instructions_sheet(wb)

    wb.save(WORKBOOK_PATH)

    # ----- sample_reviewed_decisions.csv -----
    sample_rows = build_sample_reviewed_decisions()
    write_sample_input_csv(sample_rows)

    # ----- gate を sample に対して実行し、出力を sample_ prefix でコピー -----
    gate_manifest = run_gate_on_sample(SAMPLE_INPUT_CSV, OUT_DIR)

    # ----- legal_clearance_rules.md -----
    write_legal_clearance_rules(OUT_DIR)

    # ----- manifest -----
    manifest = {
        "version": PACKET_VERSION,
        "base_commit": ALPHA_5_11_BASE_COMMIT,
        "input_packet_version": INPUT_PACKET_VERSION,
        "input_gate_version": INPUT_GATE_VERSION,
        "generated_at": GENERATED_AT,
        "generator_script": "scripts/generate_alpha5_12_reviewer_workflow_hardening.py",
        "scope": "internal_only",
        "purpose": "alpha.5.11 reviewer workbook を運用面で強化（legal_review_clearance / 拡張 export / sample fixture / 高リスク decision 受け入れ）",
        "public_release": False,
        "checked_promotion": False,
        "master_auto_update": False,
        "r8_provisional_used_for_checked": False,
        "release_pack_modified": False,
        "alpha_5_11_workbook_modified": False,
        "alpha_5_10_gate_files_modified": False,
        "alpha_5_9_packet_files_modified": False,
        "extended_workbook_columns": EXTENDED_COLUMNS,
        "valid_legal_review_clearance": list(VALID_LEGAL_CLEARANCE),
        "valid_reviewer_role": list(VALID_REVIEWER_ROLE),
        "valid_implementation_priority": list(VALID_IMPL_PRIORITY),
        "valid_implementation_risk_acknowledged": list(VALID_RISK_ACK),
        "high_risk_decisions": sorted(HIGH_RISK_DECISIONS),
        "sample_reviewed_decision_count": len(sample_rows),
        "sample_gate_result": {
            "approved_count": gate_manifest["approved_count"],
            "blocked_count": gate_manifest["blocked_count"],
            "pending_count": gate_manifest["pending_count"],
            "legal_review_required_count": gate_manifest["legal_review_required_count"],
            "future_candidate_count": gate_manifest["future_candidate_count"],
            "input_csv_schema": gate_manifest.get("input_csv_schema"),
        },
        "files": [
            "README.md",
            "alpha5_12_reviewer_decision_workbook.xlsx",
            "sample_reviewed_decisions.csv",
            "sample_reviewed_decision_validation_report.md",
            "sample_approved_changes_preview.csv",
            "sample_blocked_or_incomplete_decisions.csv",
            "sample_pending_decisions.csv",
            "sample_legal_review_required.csv",
            "legal_clearance_rules.md",
            "alpha5_12_reviewer_workflow_hardening_manifest.json",
        ],
        "invariants": [
            "master JSON 自動修正なし",
            "新規 checked 昇格なし",
            "R8.6 案資料を checked 昇格に使わない",
            "public release pack 未変更",
            "alpha.5.9 packet / alpha.5.10 gate / alpha.5.11 workbook 未破壊",
            "checked 20件 維持",
            "reviewer 入力ファイルを public に出さない",
            "後方互換: legacy 9列CSV も従来どおり読める",
            "future_candidate_only は legal clearance があっても approved にしない",
            "high risk decision は implementation_risk_acknowledged=yes が必須（拡張列ある場合）",
        ],
    }
    (OUT_DIR / "alpha5_12_reviewer_workflow_hardening_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    write_readme(OUT_DIR, manifest, gate_manifest)

    print(f"alpha.5.12 reviewer workflow hardening generated at {OUT_DIR}")
    print(f"  workbook: {WORKBOOK_PATH.name}")
    print(f"  sample_reviewed_decisions: {len(sample_rows)} rows")
    print(f"  gate result -- approved: {gate_manifest['approved_count']} / "
          f"blocked: {gate_manifest['blocked_count']} / "
          f"pending: {gate_manifest['pending_count']} / "
          f"legal_review_required: {gate_manifest['legal_review_required_count']}")


if __name__ == "__main__":
    main()
